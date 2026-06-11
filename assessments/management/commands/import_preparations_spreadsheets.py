"""Batch import English and Maths assessment spreadsheets safely.

This command imports from the school's historic spreadsheet format while keeping
existing assessment marks untouched.

Default behavior:
- Framework: "Preparations Curriculum"
- Imports configured Maths + English sheets
- Creates missing areas / sub-areas / statements as needed
- Creates new AssessmentRecord rows only when no existing row for
  (student, statement) exists
- Uses today's date unless --date is supplied
- Leaves assessed_by blank and sets notes to "Imported."

Usage (dry run):
    python manage.py import_preparations_spreadsheets

Usage (commit):
    python manage.py import_preparations_spreadsheets --commit
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from assessments.models import (
    AssessmentArea,
    AssessmentFramework,
    AssessmentRecord,
    AssessmentStatement,
    SubArea,
)
from students.models import Student, Subject


STATUS_MAP = {
    "secure": "SEC",
    "developing": "DEV",
    "emerging": "EME",
}


NAME_ALIASES = {
    "abdul rahma narma": "Abdul Rahman Narma",
    "kian linguard": "Kian Holyoake",
    "azaih hayes": "Aziah Hayes",
    "anastzja bieszczad": "Anastazja Bieszczad",
    "cian o'neil": "Cian O'Neill",
    "isaac machdo": "Isaac Machado",
    "finlay warren": "Finley Warren",
    "katie rowbottom-driver": "Katie Rowbotham-Driver",
    "ella autumn davidson": "Ella-Autumn Davinson",
    "lilly cross": "Lilly-Ann Cross",
    "muhammad seedat": "Muhammad Yusuf Seedat",
    "alojzy wrotnicki": "Alojzy-Rudy Wrotnicki",
    "adam akyildis": "Adam Akyildiz",
    "alfie - joe walmsley": "Alfie Walmsley",
    "asad bhatti": "Abdulrahman Bhatti",
    "dani - kaylen khan": "Mohammad Khan",
    "devanah mistry": "Devansh Mistry",
    "dhruv parkeh": "Dhruv Parekh",
    "fadumsahra hared": "Fadumasahra Hared",
    "ishan miah": "Minhaj Miah",
    "john george mcginley": "John McGinlay",
    "kieran nutting": "Kirean Nutting",
    "mannan mehul": "Maanan Mehul",
    "mohammad hashi dar": "Muhammad Dar",
    "muhammad rostram": "Muhammad Rostam",
    "olamide orungheia": "Olamide Orungbeja",
    "onosetale ehika": "Onosetale Jude",
    "scarlet pochin": "Scarlett Pochin",
    "taydn murrel": "Taydn Murrell",
    "timmy szabo": "Timothy Szabo",
    "yariette bangoura": "Fatoumata Bangoura",
    "yasin moursey": "Yasin Moursy",
}

DEFAULT_ENGLISH_SHEETS = [
    "Reading - Word Reading",
    "Reading - Comprehension",
    "Writing - Handwriting",
    "Writing - V, G and P",
    "Writing - Composition",
    "Writing - Transcription",
    "Entry Level",
]


DEFAULT_MATHS_SHEETS = [
    ("Topic 0 - Futures", "Maths Topic 0 - Entry and Foundations"),
    ("Topic 0 - Preparations", "Maths Topic 0 - Entry and Foundations"),
    (
        "Topic 1 - Futures ",
        "Maths Topic 1 - Using Numbers and the Number System - Whole Numbers, Fractions and Decimals",
    ),
    (
        "Topic 1 - Preparations",
        "Maths Topic 1 - Using Numbers and the Number System - Whole Numbers, Fractions and Decimals",
    ),
    (
        "Topic 2 - Futures",
        "Maths Topic 2 - Using Common Measures, Shape and Space",
    ),
    (
        "Topic 2 - Preparations",
        "Maths Topic 2 - Using Common Measures, Shape and Space",
    ),
    (
        "Topic 3 - Futures",
        "Maths Topic 3 - Handling Data and Information",
    ),
    (
        "Topic 3 - Preparations",
        "Maths Topic 3 - Handling Data and Information",
    ),
    ("Entry Level", "Maths Entry Level"),
]


def _normalise(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"\s*-\s*", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


NORMALISED_NAME_ALIASES = {
    _normalise(alias): target for alias, target in NAME_ALIASES.items()
}


class Command(BaseCommand):
    help = (
        "Import English + Maths spreadsheets into Preparations Curriculum "
        "without overwriting existing marks."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--english-file",
            default="Input/English Assessment.xlsx",
            help="Path to English spreadsheet.",
        )
        parser.add_argument(
            "--maths-file",
            default="Input/Maths Assessment .xlsx",
            help="Path to Maths spreadsheet.",
        )
        parser.add_argument(
            "--framework",
            default="Preparations Curriculum",
            help="Framework name.",
        )
        parser.add_argument(
            "--english-subject",
            default="English",
            help="Subject name for English sheets.",
        )
        parser.add_argument(
            "--maths-subject",
            default="Maths",
            help="Subject name for Maths sheets.",
        )
        parser.add_argument(
            "--date",
            default=dt.date.today().isoformat(),
            help="Assessed date in YYYY-MM-DD format. Defaults to today.",
        )
        parser.add_argument(
            "--notes",
            default="Imported.",
            help="Notes value for imported records.",
        )
        parser.add_argument(
            "--commit",
            action="store_true",
            help="Write to DB. Without this flag, runs as dry-run.",
        )

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is not installed.") from exc

        try:
            assessed_date = dt.date.fromisoformat(opts["date"])
        except ValueError as exc:
            raise CommandError(
                f"Bad --date: {opts['date']} (expected YYYY-MM-DD)"
            ) from exc

        english_path = Path(opts["english_file"])
        maths_path = Path(opts["maths_file"])
        if not english_path.exists():
            raise CommandError(f"English file not found: {english_path}")
        if not maths_path.exists():
            raise CommandError(f"Maths file not found: {maths_path}")

        commit = opts["commit"]
        mode = "COMMIT" if commit else "DRY-RUN"

        self.stdout.write(self.style.NOTICE(f"=== Mode: {mode} ==="))
        self.stdout.write(f"Framework: {opts['framework']}")
        self.stdout.write(f"Date:      {assessed_date}")
        self.stdout.write(f"Notes:     {opts['notes']!r}")
        self.stdout.write("")

        framework, _ = AssessmentFramework.objects.get_or_create(
            name=opts["framework"],
            defaults={"description": "Imported historic marks.", "is_active": True},
        )

        english_subject, eng_created = Subject.objects.get_or_create(
            name=opts["english_subject"],
            defaults={
                "applicable_pathways": ["PREP"],
                "applicable_phases": [1, 2],
                "is_active": True,
            },
        )
        if eng_created:
            self.stdout.write(self.style.SUCCESS(
                f"Created subject: {english_subject.name}"
            ))

        maths_subject, _ = Subject.objects.get_or_create(
            name=opts["maths_subject"],
            defaults={
                "applicable_pathways": ["PREP"],
                "applicable_phases": [1, 2],
                "is_active": True,
            },
        )

        student_index = {}
        for student in Student.objects.filter(is_active=True):
            full_name = f"{student.first_name} {student.last_name}".strip()
            student_index[_normalise(full_name)] = student

        jobs = []
        for sheet_name in DEFAULT_ENGLISH_SHEETS:
            jobs.append((english_path, sheet_name, english_subject, sheet_name))
        for sheet_name, area_name in DEFAULT_MATHS_SHEETS:
            jobs.append((maths_path, sheet_name, maths_subject, area_name))

        summary = {
            "sheets_processed": 0,
            "areas_created": 0,
            "sub_areas_created": 0,
            "statements_created": 0,
            "records_created": 0,
            "records_skipped_existing": 0,
            "records_skipped_unknown_status": 0,
            "records_considered": 0,
        }
        unmatched_by_sheet = {}

        wb_cache = {}

        def get_workbook(path_obj):
            cached = wb_cache.get(path_obj)
            if cached is None:
                cached = openpyxl.load_workbook(path_obj, data_only=True, read_only=True)
                wb_cache[path_obj] = cached
            return cached

        try:
            with transaction.atomic():
                for path_obj, sheet_name, subject, area_name in jobs:
                    workbook = get_workbook(path_obj)
                    if sheet_name not in workbook.sheetnames:
                        self.stdout.write(
                            self.style.WARNING(
                                f"Skip missing sheet {sheet_name!r} in {path_obj.name}"
                            )
                        )
                        continue

                    sheet_summary = self._import_sheet(
                        ws=workbook[sheet_name],
                        framework=framework,
                        subject=subject,
                        area_name=area_name,
                        assessed_date=assessed_date,
                        notes=opts["notes"],
                        student_index=student_index,
                        unmatched_by_sheet=unmatched_by_sheet,
                    )
                    summary["sheets_processed"] += 1
                    for key, value in sheet_summary.items():
                        summary[key] += value

                if not commit:
                    raise RuntimeError("dry-run rollback")
        except RuntimeError as exc:
            if str(exc) != "dry-run rollback":
                raise
        finally:
            for wb in wb_cache.values():
                wb.close()

        self.stdout.write("")
        self.stdout.write(self.style.NOTICE("Summary"))
        self.stdout.write(f"  Sheets processed:            {summary['sheets_processed']}")
        self.stdout.write(f"  Areas created:              {summary['areas_created']}")
        self.stdout.write(f"  Sub-areas created:          {summary['sub_areas_created']}")
        self.stdout.write(f"  Statements created:         {summary['statements_created']}")
        self.stdout.write(f"  Marks considered:           {summary['records_considered']}")
        self.stdout.write(f"  Records created:            {summary['records_created']}")
        self.stdout.write(f"  Skipped existing marks:     {summary['records_skipped_existing']}")
        self.stdout.write(f"  Skipped unknown statuses:   {summary['records_skipped_unknown_status']}")

        if unmatched_by_sheet:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Unmatched students"))
            for sheet_name in sorted(unmatched_by_sheet.keys()):
                names = sorted(unmatched_by_sheet[sheet_name])
                self.stdout.write(f"  {sheet_name}:")
                for name in names:
                    self.stdout.write(f"    - {name}")

        self.stdout.write("")
        if commit:
            self.stdout.write(self.style.SUCCESS("COMMIT complete."))
        else:
            self.stdout.write(self.style.NOTICE("DRY-RUN complete (no data written)."))

    def _import_sheet(
        self,
        ws,
        framework,
        subject,
        area_name,
        assessed_date,
        notes,
        student_index,
        unmatched_by_sheet,
    ):
        rows_iter = ws.iter_rows(values_only=True)
        try:
            row1 = next(rows_iter)
            row2 = next(rows_iter)
        except StopIteration:
            return {
                "areas_created": 0,
                "sub_areas_created": 0,
                "statements_created": 0,
                "records_created": 0,
                "records_skipped_existing": 0,
                "records_skipped_unknown_status": 0,
                "records_considered": 0,
            }

        area, area_created = AssessmentArea.objects.get_or_create(
            framework=framework,
            subject=subject,
            name=area_name,
            defaults={"phase": None, "year_group": None},
        )

        column_specs = []
        current_level = None
        max_cols = max(len(row1 or ()), len(row2 or ()))
        for col_idx in range(2, max_cols):
            level_cell = row1[col_idx] if col_idx < len(row1) else None
            if level_cell is not None and str(level_cell).strip():
                current_level = str(level_cell).strip()

            text_cell = row2[col_idx] if col_idx < len(row2) else None
            if text_cell is None:
                continue
            text = str(text_cell).replace("\n", " ").strip()
            text = " ".join(text.split())
            if not text:
                continue
            level = current_level or "(no level)"
            column_specs.append((col_idx, level, text))

        sub_area_map = {}
        sub_areas_created = 0
        for order, level in enumerate(dict.fromkeys(level for _, level, _ in column_specs)):
            sub_area, created = SubArea.objects.get_or_create(
                area=area,
                name=level,
                defaults={"order": order},
            )
            if created:
                sub_areas_created += 1
            sub_area_map[level] = sub_area

        statement_map = {}
        statements_created = 0
        for order, (col_idx, level, text) in enumerate(column_specs):
            statement, created = AssessmentStatement.objects.get_or_create(
                area=area,
                sub_area=sub_area_map[level],
                statement_text=text,
                defaults={"order": order},
            )
            if created:
                statements_created += 1
            statement_map[col_idx] = statement

        records_created = 0
        records_skipped_existing = 0
        records_skipped_unknown_status = 0
        records_considered = 0

        for row in rows_iter:
            if not row:
                continue

            name_cell = row[0] if len(row) > 0 else None
            raw_name = str(name_cell).strip() if name_cell else ""
            if not raw_name:
                continue

            student = student_index.get(_normalise(raw_name))
            if student is None:
                alias_target = NORMALISED_NAME_ALIASES.get(_normalise(raw_name))
                if alias_target:
                    student = student_index.get(_normalise(alias_target))

            if student is None:
                unmatched_by_sheet.setdefault(ws.title, set()).add(raw_name)
                continue

            for col_idx, statement in statement_map.items():
                if col_idx >= len(row):
                    continue
                mark_val = row[col_idx]
                if mark_val is None or not str(mark_val).strip():
                    continue

                records_considered += 1
                status_code = STATUS_MAP.get(str(mark_val).strip().lower())
                if not status_code:
                    records_skipped_unknown_status += 1
                    continue

                if AssessmentRecord.objects.filter(
                    student=student, statement=statement
                ).exists():
                    records_skipped_existing += 1
                    continue

                AssessmentRecord.objects.create(
                    student=student,
                    statement=statement,
                    status=status_code,
                    assessed_by=None,
                    assessed_date=assessed_date,
                    notes=notes,
                )
                records_created += 1

        return {
            "areas_created": 1 if area_created else 0,
            "sub_areas_created": sub_areas_created,
            "statements_created": statements_created,
            "records_created": records_created,
            "records_skipped_existing": records_skipped_existing,
            "records_skipped_unknown_status": records_skipped_unknown_status,
            "records_considered": records_considered,
        }