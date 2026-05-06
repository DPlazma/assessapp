"""Import historic Maths/English marks from a wide-format spreadsheet.

Usage (dry run by default — no DB writes):
    python manage.py import_historic_marks <file.xlsx> \
        --sheet "Topic 1 - Preparations" \
        --framework "Preparations Curriculum" \
        --subject "Maths" \
        --area "Maths — Topic 1" \
        --phase 1 \
        --date 2026-05-05

Add --commit to actually write.

Spreadsheet shape:
    Row 1: level headers (sparse, e.g. 'Foundation Level', 'Level 1' ...)
    Row 2: statement texts (181-ish columns), col A blank, col B = 'Current Level'
    Row 3+: col A = student name, col B = current-level summary (ignored),
            cols C+ = mark values ('Secure' / 'Developing' / 'Emerging' or blank)
    Blank rows in col A are group separators and are skipped.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


STATUS_MAP = {
    "secure": "SEC",
    "developing": "DEV",
    "emerging": "EME",
}

# Manual aliases: spreadsheet name (lower-cased, hyphens/spaces ignored) → canonical "First Last"
# in the database. Used when the spreadsheet contains a typo or a different given/legal name.
NAME_ALIASES = {
    "abdul rahma narma": "Abdul Rahman Narma",
    "kian linguard": "Kian Holyoake",
    "azaih hayes": "Aziah Hayes",
    "anastzja bieszczad": "Anastazja Bieszczad",
    "cian o'neil": "Cian O'Neill",
    "isaac machdo": "Isaac Machado",
    "finlay warren": "Finley Warren",
    "katie rowbottom-driver": "Katie Rowbotham-Driver",
    "ella autumn davidson": "Ella-Autumn Davinson",
    "lilly cross": "Lilly-Ann Cross",
    "muhammad seedat": "Muhammad Yusuf Seedat",
    "alojzy wrotnicki": "Alojzy-Rudy Wrotnicki",
}


def _normalise(name: str) -> str:
    """Lower-case, collapse spaces, treat ' - ' / '-' / spaces around hyphens as the same.

    'Jacob James - Hill' and 'Jacob-James Hill' both normalise to 'jacob james hill'."""
    s = name.strip().lower()
    # Replace any hyphen (with optional surrounding whitespace) with a single space.
    import re as _re
    s = _re.sub(r"\s*-\s*", " ", s)
    s = _re.sub(r"\s+", " ", s)
    return s


class Command(BaseCommand):
    help = "Import historic assessment marks from a wide-format XLSX."

    def add_arguments(self, parser):
        parser.add_argument("file", help="Path to the .xlsx file.")
        parser.add_argument("--sheet", required=True, help="Sheet name to import.")
        parser.add_argument("--framework", required=True, help="Framework name (created if missing).")
        parser.add_argument("--subject", required=True, help="Subject name (created if missing).")
        parser.add_argument("--area", required=True, help="Area name within the framework (created if missing).")
        parser.add_argument("--phase", type=int, default=None, help="Optional phase (1 or 2).")
        parser.add_argument("--year-group", type=int, default=None, help="Optional year group for the area.")
        parser.add_argument("--date", required=True, help="Assessed date in YYYY-MM-DD format.")
        parser.add_argument("--pathway", default="PREP",
                            help="Comma-separated pathway codes for the subject (default: PREP).")
        parser.add_argument("--commit", action="store_true",
                            help="Actually write to the database. Default is dry-run.")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl is not installed.")

        from assessments.models import (
            AssessmentFramework, AssessmentArea, AssessmentStatement,
            AssessmentRecord, SubArea,
        )
        from students.models import Student, Subject

        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        try:
            assessed_date = dt.date.fromisoformat(opts["date"])
        except ValueError:
            raise CommandError(f"Bad --date: {opts['date']} (expected YYYY-MM-DD)")

        commit = opts["commit"]
        mode_label = "COMMIT" if commit else "DRY-RUN"

        self.stdout.write(self.style.NOTICE(f"=== Mode: {mode_label} ==="))
        self.stdout.write(f"File:      {path}")
        self.stdout.write(f"Sheet:     {opts['sheet']}")
        self.stdout.write(f"Framework: {opts['framework']}")
        self.stdout.write(f"Subject:   {opts['subject']}")
        self.stdout.write(f"Area:      {opts['area']}  phase={opts['phase']}  year={opts['year_group']}")
        self.stdout.write(f"Date:      {assessed_date}")
        self.stdout.write("")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if opts["sheet"] not in wb.sheetnames:
            raise CommandError(
                f"Sheet {opts['sheet']!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[opts["sheet"]]

        # Read row 1 (level headers, sparse) and row 2 (statement texts)
        rows_iter = ws.iter_rows(values_only=True)
        row1 = next(rows_iter)
        row2 = next(rows_iter)

        # Build column → (level, statement_text) map starting from col C (index 2)
        # Forward-fill the level header across blank cells.
        column_specs = []  # list of (col_index, level_label, statement_text)
        current_level = None
        for col_idx in range(2, len(row2)):
            level_cell = row1[col_idx] if col_idx < len(row1) else None
            if level_cell is not None and str(level_cell).strip():
                current_level = str(level_cell).strip()
            text_cell = row2[col_idx] if col_idx < len(row2) else None
            if text_cell is None:
                continue
            text = str(text_cell).replace("\n", " ").strip()
            text = " ".join(text.split())  # collapse internal whitespace
            if not text:
                continue
            column_specs.append((col_idx, current_level or "(no level)", text))

        self.stdout.write(f"Found {len(column_specs)} statement columns across "
                          f"{len(set(c[1] for c in column_specs))} levels.")

        # Read body rows: col A = student name, marks in col C+
        student_rows = []  # list of (row_idx, name, {col_idx: value})
        for row_idx, row in enumerate(rows_iter, start=3):
            if not row:
                continue
            name_cell = row[0] if len(row) > 0 else None
            name = (str(name_cell).strip() if name_cell else "")
            if not name:
                continue  # blank separator row
            marks = {}
            for col_idx, _level, _text in column_specs:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None and str(val).strip():
                        marks[col_idx] = str(val).strip()
            student_rows.append((row_idx, name, marks))

        self.stdout.write(f"Found {len(student_rows)} student rows.")
        self.stdout.write("")

        # ── Match students ──────────────────────────────────────────
        # Build two indexes on the DB:
        #   1) normalised "first last" → Student
        #   2) same, but also keyed by the canonical alias target
        all_students = list(Student.objects.filter(is_active=True))
        name_index = {}
        for s in all_students:
            full = f"{s.first_name} {s.last_name}".strip()
            name_index[_normalise(full)] = s

        matched, unmatched, alias_used = [], [], []
        for row_idx, name, marks in student_rows:
            key = _normalise(name)
            student = name_index.get(key)
            if student is None:
                # Try alias map (normalise alias keys to match)
                alias_target = None
                for ak, av in NAME_ALIASES.items():
                    if _normalise(ak) == key:
                        alias_target = av
                        break
                if alias_target:
                    student = name_index.get(_normalise(alias_target))
                    if student:
                        alias_used.append((name, alias_target))
            if student:
                matched.append((row_idx, name, student, marks))
            else:
                unmatched.append((row_idx, name, marks))

        self.stdout.write(self.style.SUCCESS(f"Matched students:   {len(matched)}"))
        if alias_used:
            self.stdout.write(f"  ({len(alias_used)} via alias / hyphen-normalisation)")
            for spreadsheet, target in alias_used:
                self.stdout.write(f"    · {spreadsheet!r} → {target!r}")
        if unmatched:
            self.stdout.write(self.style.WARNING(f"Unmatched students: {len(unmatched)}"))
            for row_idx, name, marks in unmatched:
                self.stdout.write(f"  · row {row_idx}: {name!r}  ({len(marks)} marks would be skipped)")
        self.stdout.write("")

        # ── Validate mark values ────────────────────────────────────
        bad_values = set()
        total_marks = 0
        for _r, _n, _s, marks in matched:
            for v in marks.values():
                if v.lower() not in STATUS_MAP:
                    bad_values.add(v)
                else:
                    total_marks += 1
        if bad_values:
            self.stdout.write(self.style.ERROR(
                f"Unknown mark values found (will be skipped): {sorted(bad_values)}"
            ))
        self.stdout.write(f"Total marks ready to import: {total_marks}")
        self.stdout.write("")

        # ── Plan the writes ─────────────────────────────────────────
        levels_in_data = []
        seen_levels = set()
        for _c, level, _t in column_specs:
            if level not in seen_levels:
                seen_levels.add(level)
                levels_in_data.append(level)

        self.stdout.write("Plan:")
        self.stdout.write(f"  · Framework:   {opts['framework']}  (created if missing)")
        self.stdout.write(f"  · Subject:     {opts['subject']}     (created if missing, pathway={opts['pathway']})")
        self.stdout.write(f"  · Area:        {opts['area']}        (created if missing)")
        for lvl in levels_in_data:
            n_stmt = sum(1 for c in column_specs if c[1] == lvl)
            self.stdout.write(f"      ├─ Sub-area {lvl!r:25s}  ({n_stmt} statements)")
        self.stdout.write(f"  · AssessmentRecords to write: {total_marks}")
        self.stdout.write("")

        if not commit:
            self.stdout.write(self.style.NOTICE(
                "DRY-RUN complete — re-run with --commit to apply."
            ))
            return

        # ── COMMIT path ─────────────────────────────────────────────
        with transaction.atomic():
            framework, _ = AssessmentFramework.objects.get_or_create(
                name=opts["framework"],
                defaults={"description": "Imported via import_historic_marks."},
            )

            pathway_codes = [p.strip() for p in opts["pathway"].split(",") if p.strip()]
            subject, sub_created = Subject.objects.get_or_create(
                name=opts["subject"],
                defaults={
                    "applicable_pathways": pathway_codes,
                    "applicable_phases": [1, 2],
                    "is_active": True,
                },
            )
            # Don't overwrite an existing subject's pathways silently.
            if not sub_created:
                self.stdout.write(f"  Subject {subject.name!r} already exists — leaving its pathway/phase config untouched.")

            area, area_created = AssessmentArea.objects.get_or_create(
                framework=framework,
                subject=subject,
                name=opts["area"],
                defaults={
                    "phase": opts["phase"],
                    "year_group": opts["year_group"],
                },
            )

            # Sub-areas + statements
            sub_area_map = {}  # level_label → SubArea
            for order, lvl in enumerate(levels_in_data):
                sa, _ = SubArea.objects.get_or_create(
                    area=area, name=lvl,
                    defaults={"order": order},
                )
                sub_area_map[lvl] = sa

            stmt_map = {}  # col_idx → AssessmentStatement
            for order, (col_idx, level, text) in enumerate(column_specs):
                sa = sub_area_map[level]
                stmt, _ = AssessmentStatement.objects.get_or_create(
                    area=area, sub_area=sa, statement_text=text,
                    defaults={"order": order},
                )
                stmt_map[col_idx] = stmt

            # Records
            written = 0
            for _row_idx, _name, student, marks in matched:
                for col_idx, val in marks.items():
                    code = STATUS_MAP.get(val.lower())
                    if not code:
                        continue
                    stmt = stmt_map.get(col_idx)
                    if not stmt:
                        continue
                    AssessmentRecord.objects.create(
                        student=student,
                        statement=stmt,
                        status=code,
                        assessed_by=None,
                        assessed_date=assessed_date,
                        notes="Imported from historic spreadsheet.",
                    )
                    written += 1

        self.stdout.write(self.style.SUCCESS(
            f"COMMIT complete — wrote {written} AssessmentRecord rows."
        ))
