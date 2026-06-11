import csv
import io
import json

from django.db import models
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_POST

from core.decorators import slt_or_subject_lead_required
from students.models import Student, Subject, ClassGroup
from staff.models import ClassAssignment, ClassCover, SubjectLead
from core.models import Term, AcademicYear
from .models import (
    AssessmentFramework,
    AssessmentArea,
    SubArea,
    AssessmentStatement,
    AssessmentRecord,
    AssessmentSnapshot,
    FrameworkAssignment,
    PersonalisedFramework,
)

MAPP_FRAMEWORK_NAME = "MAPP (Mapping and Assessing Personal Progress)"
STATUS_VALUE = {"NYA": 0, "EME": 1, "DEV": 2, "SEC": 3}


def _can_assess_student(user, student, subject=None):
    """Check if user can record assessments for this student.

    Args:
        subject: Optional Subject instance. When provided, also checks
                 whether the user is a subject lead for that subject
                 (subject leads can assess any student in their subject).
    """
    if not hasattr(user, "staffprofile"):
        return False
    profile = user.staffprofile
    if profile.is_slt:
        return True
    # Subject leads can assess any student in their subject(s)
    if subject and profile.is_subject_lead:
        if SubjectLead.objects.filter(user=user, subject=subject).exists():
            return True
    today = timezone.now().date()
    # Check designated classes
    assigned = ClassAssignment.objects.filter(
        user=user, class_group=student.class_group
    ).exists()
    if assigned:
        return True
    # Check active covers
    covered = ClassCover.objects.filter(
        user=user,
        class_group=student.class_group,
        start_date__lte=today,
        end_date__gte=today,
    ).exists()
    return covered


@login_required
def my_subjects(request):
    """Subject lead landing page — shows their subjects grouped by pathway and class."""
    profile = getattr(request.user, "staffprofile", None)
    if not profile or not profile.is_subject_lead:
        return redirect("core:dashboard")

    from students.models import PATHWAY_CHOICES

    pathway_label = dict(PATHWAY_CHOICES)

    lead_subjects = Subject.objects.filter(
        leads__user=request.user, is_active=True
    ).order_by("order", "name")

    subject_data = []
    for subj in lead_subjects:
        students = list(
            Student.objects.filter(
                is_active=True,
                pathway__in=subj.applicable_pathways,
                phase__in=subj.applicable_phases,
            )
            .select_related("class_group")
            .order_by("pathway", "class_group__name", "last_name", "first_name")
        )

        # Group: pathway → class → [students]
        pathway_groups = []
        pathway_buckets = {}
        for s in students:
            pw_code = s.pathway or ""
            pw_bucket = pathway_buckets.get(pw_code)
            if pw_bucket is None:
                pw_bucket = {
                    "code": pw_code,
                    "label": pathway_label.get(pw_code, pw_code or "Unassigned"),
                    "classes": [],
                    "_class_index": {},
                    "count": 0,
                }
                pathway_buckets[pw_code] = pw_bucket
                pathway_groups.append(pw_bucket)

            cg = s.class_group
            cg_key = cg.pk if cg else None
            cls_bucket = pw_bucket["_class_index"].get(cg_key)
            if cls_bucket is None:
                cls_bucket = {
                    "class_group": cg,
                    "name": cg.name if cg else "No class",
                    "students": [],
                }
                pw_bucket["_class_index"][cg_key] = cls_bucket
                pw_bucket["classes"].append(cls_bucket)
            cls_bucket["students"].append(s)
            pw_bucket["count"] += 1

        # Strip the internal index before sending to the template
        for pw in pathway_groups:
            pw.pop("_class_index", None)

        subject_data.append({
            "subject": subj,
            "total": len(students),
            "pathway_groups": pathway_groups,
        })

    context = {"subject_data": subject_data}
    return render(request, "assessments/my_subjects.html", context)


@login_required
def student_subjects(request, student_id):
    """Student hub — overview of frameworks, EHCP, MAPP, and subjects."""
    student = get_object_or_404(Student, pk=student_id)

    # ── Assigned frameworks (direct + via class group) ──
    from evidence.models import MAPPConfig
    direct_assignments = FrameworkAssignment.objects.filter(
        student=student,
    ).select_related("framework")
    class_assignments = FrameworkAssignment.objects.none()
    if student.class_group:
        class_assignments = FrameworkAssignment.objects.filter(
            class_group=student.class_group,
        ).select_related("framework")

    assigned_frameworks = []
    seen_fw_ids = set()
    for a in list(direct_assignments) + list(class_assignments):
        if a.framework_id not in seen_fw_ids and a.framework.is_active:
            seen_fw_ids.add(a.framework_id)
            # Check if personalised
            pf = PersonalisedFramework.objects.filter(
                student=student, framework=a.framework
            ).first()
            assigned_frameworks.append({
                "framework": a.framework,
                "via": "direct" if a.student_id else "class",
                "personalised_count": pf.statements.count() if pf else None,
            })

    # ── Subjects from assigned frameworks' assessment areas ──
    subjects = Subject.objects.filter(
        is_active=True,
        assessment_areas__framework_id__in=seen_fw_ids,
    ).distinct().order_by("order", "name")

    # Check per-subject so subject leads see the right buttons
    can_assess_any = False
    subject_permissions = {}
    for subj in subjects:
        perm = _can_assess_student(request.user, student, subj)
        subject_permissions[subj.pk] = perm
        if perm:
            can_assess_any = True

    # ── EHCP targets summary ──
    from evidence.models import EHCPTarget
    ehcp_targets = EHCPTarget.objects.filter(student=student).order_by("-created_at")[:5]
    ehcp_total = EHCPTarget.objects.filter(student=student).count()

    # ── MAPP summary (only if MAPP framework is assigned to this student) ──
    mapp_config = None
    mapp_priorities = []
    has_mapp = False
    _mapp_cfg = MAPPConfig.get_active()
    if _mapp_cfg:
        # Check if the MAPP framework is actually assigned to this student
        from assessments.views import MAPP_FRAMEWORK_NAME
        mapp_fw_ids = set(
            AssessmentFramework.objects.filter(
                name=MAPP_FRAMEWORK_NAME, is_active=True
            ).values_list("pk", flat=True)
        )
        if mapp_fw_ids & seen_fw_ids:
            has_mapp = True
            mapp_config = _mapp_cfg
            from evidence.models import MAPPLearningPriority
            mapp_priorities = MAPPLearningPriority.objects.filter(
                student=student
            ).order_by("-created_at")[:3]

    # ── Interventions count ──
    from evidence.models import InterventionEnrolment
    intervention_count = InterventionEnrolment.objects.filter(
        student=student, status="ACTIVE"
    ).count()

    context = {
        "student": student,
        "subjects": subjects,
        "can_assess": can_assess_any,
        "subject_permissions": subject_permissions,
        "assigned_frameworks": assigned_frameworks,
        "ehcp_targets": ehcp_targets,
        "ehcp_total": ehcp_total,
        "has_ehcp": ehcp_total > 0,
        "has_mapp": has_mapp,
        "mapp_config": mapp_config,
        "mapp_priorities": mapp_priorities,
        "has_interventions": intervention_count > 0,
        "intervention_count": intervention_count,
    }
    return render(request, "assessments/student_subjects.html", context)


@login_required
def assess_student(request, student_id, subject_id):
    """Show statements for a student+subject, respecting framework assignments
    and per-student personalisation."""
    student = get_object_or_404(Student, pk=student_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    can_assess = _can_assess_student(request.user, student, subject)

    # ── Determine which frameworks are assigned to this student ──
    # Direct student assignments
    direct_fw_ids = set(
        FrameworkAssignment.objects.filter(
            student=student, framework__is_active=True
        ).values_list("framework_id", flat=True)
    )
    # Via class group
    class_fw_ids = set()
    if student.class_group:
        class_fw_ids = set(
            FrameworkAssignment.objects.filter(
                class_group=student.class_group, framework__is_active=True
            ).values_list("framework_id", flat=True)
        )
    assigned_fw_ids = direct_fw_ids | class_fw_ids

    # If the student has framework assignments, filter areas to those frameworks.
    # Otherwise fall back to showing everything (backward-compatible).
    if assigned_fw_ids:
        areas = AssessmentArea.objects.filter(
            subject=subject, framework_id__in=assigned_fw_ids
        ).prefetch_related("statements", "sub_areas__statements")
        frameworks = AssessmentFramework.objects.filter(
            pk__in=assigned_fw_ids, areas__subject=subject
        ).distinct()
    else:
        areas = AssessmentArea.objects.filter(subject=subject).prefetch_related(
            "statements", "sub_areas__statements"
        )
        frameworks = AssessmentFramework.objects.filter(
            is_active=True, areas__subject=subject
        ).distinct()

    # Optional framework filter from query string
    framework_id = request.GET.get("framework")
    if framework_id:
        areas = areas.filter(framework_id=framework_id)

    # ── Per-student personalisation ──
    # Look up which statements have been personalised for this student.
    personalised_stmt_ids = None  # None means "show everything"
    pf_qs = PersonalisedFramework.objects.filter(student=student)
    if framework_id:
        pf_qs = pf_qs.filter(framework_id=framework_id)
    pf_list = list(pf_qs.prefetch_related("statements"))
    if pf_list:
        personalised_stmt_ids = set()
        for pf in pf_list:
            stmt_ids = set(pf.statements.values_list("pk", flat=True))
            if stmt_ids:
                personalised_stmt_ids |= stmt_ids
        # If all personalisations had empty selections, treat as "show all"
        if not personalised_stmt_ids:
            personalised_stmt_ids = None

    # Get latest assessment for each visible statement
    areas = list(areas)  # evaluate once for stats + template reuse
    assessment_map = {}
    for area in areas:
        for statement in area.statements.all():
            if personalised_stmt_ids is not None and statement.pk not in personalised_stmt_ids:
                continue
            latest = (
                AssessmentRecord.objects.filter(student=student, statement=statement)
                .order_by("-assessed_date", "-created_at")
                .first()
            )
            assessment_map[statement.pk] = latest

    # ── Compute NEDS counters per area and sub-area ──
    for area in areas:
        counts = {'NYA': 0, 'EME': 0, 'DEV': 0, 'SEC': 0, 'total': 0}
        for stmt in area.statements.all():
            if personalised_stmt_ids is not None and stmt.pk not in personalised_stmt_ids:
                continue
            counts['total'] += 1
            rec = assessment_map.get(stmt.pk)
            st = rec.status if rec else 'NYA'
            counts[st] += 1
        area.neds_n = counts['NYA']
        area.neds_e = counts['EME']
        area.neds_d = counts['DEV']
        area.neds_s = counts['SEC']
        area.neds_total = counts['total']
        for sub in area.sub_areas.all():
            sc = {'NYA': 0, 'EME': 0, 'DEV': 0, 'SEC': 0, 'total': 0}
            for stmt in sub.statements.all():
                if personalised_stmt_ids is not None and stmt.pk not in personalised_stmt_ids:
                    continue
                sc['total'] += 1
                rec = assessment_map.get(stmt.pk)
                st = rec.status if rec else 'NYA'
                sc[st] += 1
            sub.neds_n = sc['NYA']
            sub.neds_e = sc['EME']
            sub.neds_d = sc['DEV']
            sub.neds_s = sc['SEC']
            sub.neds_total = sc['total']

    # AI options context
    from evidence.models import EHCPOutcome
    cohort_options = [
        ("eyfs", "EYFS / Early Years"),
        ("sensory", "Sensory learners"),
        ("nonverbal", "Non-verbal / pre-verbal"),
        ("pmld", "PMLD"),
        ("preformal", "Pre-formal curriculum"),
        ("semiformal", "Semi-formal curriculum"),
        ("formal", "Formal curriculum"),
        ("asc", "Autism / ASC"),
        ("physical", "Physical / motor"),
    ]

    context = {
        "student": student,
        "subject": subject,
        "areas": areas,
        "assessment_map": assessment_map,
        "can_assess": can_assess,
        "frameworks": frameworks,
        "selected_framework": framework_id,
        "status_choices": AssessmentRecord.STATUS_CHOICES,
        "personalised_stmt_ids": personalised_stmt_ids,
        "ai_available": _ai_available(),
        "cohort_options": cohort_options,
        "ehcp_outcomes": EHCPOutcome.objects.all(),
    }
    return render(request, "assessments/assess_student.html", context)


def _ai_available():
    from core.models import AISettings
    ai = AISettings.load()
    return ai.enabled and ai.provider != "none"


@login_required
@require_POST
def ai_next_steps(request, student_id, subject_id):
    """POST endpoint: AI suggests next teaching steps with structured options.

    Expects JSON body with:
      - cohorts: list of cohort tags  (e.g. ["sensory", "nonverbal"])
      - mode: "targets" | "steps_only" | "strategies"
      - outcome_id: optional EHCP outcome PK
    Returns JSON {suggestions: [...], mode: str} or {error: str}.
    """
    from core.ai import ai_chat

    student = get_object_or_404(Student, pk=student_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    try:
        body = json.loads(request.body)
    except (ValueError, TypeError):
        body = {}

    cohorts = body.get("cohorts", [])
    mode = body.get("mode", "targets")

    # ── Build assessment data context ──
    areas = AssessmentArea.objects.filter(subject=subject).prefetch_related("statements")
    lines = []
    framework_names = set()
    for area in areas:
        framework_names.add(area.framework.name)
        area_statuses = []
        for stmt in area.statements.all():
            latest = (
                AssessmentRecord.objects.filter(student=student, statement=stmt)
                .order_by("-assessed_date", "-created_at")
                .first()
            )
            status = latest.get_status_display() if latest else "Not Yet Assessed"
            area_statuses.append(f"  - {stmt.statement_text}: {status}")
        if area_statuses:
            lines.append(f"{area.name}:")
            lines.extend(area_statuses)

    data_text = "\n".join(lines)
    fw_text = ", ".join(sorted(framework_names)) if framework_names else "General"

    # ── Build cohort context ──
    cohort_labels = {
        "eyfs": "EYFS / Early Years learner",
        "sensory": "sensory-focused learner who benefits from multi-sensory approaches",
        "nonverbal": "non-verbal or pre-verbal communicator (uses AAC, symbols, signing)",
        "pmld": "learner with profound and multiple learning difficulties (PMLD)",
        "preformal": "learner on a pre-formal curriculum (engagement-level)",
        "semiformal": "learner on a semi-formal curriculum",
        "formal": "learner on a formal / modified National Curriculum pathway",
        "asc": "autistic learner who benefits from structured, predictable approaches",
        "physical": "learner with physical or motor difficulties",
    }
    cohort_text = ""
    if cohorts:
        descs = [cohort_labels.get(c, c) for c in cohorts if c in cohort_labels]
        if descs:
            cohort_text = (
                f"\nLearner profile: This is a {', and a '.join(descs)}. "
                "Tailor all suggestions to be appropriate and accessible for this profile.\n"
            )

    # ── Build mode-specific prompt ──
    if mode == "targets":
        task_text = (
            "Generate 3-5 EHCP-style targets with small steps.\n"
            "For each target provide:\n"
            "- A clear, measurable target title (suitable as an EHCP target)\n"
            "- A brief rationale (1 sentence explaining why this target)\n"
            "- 3-5 small steps (lettered A, B, C…) that break the target into "
            "granular, observable milestones a teaching assistant can work on\n\n"
            "Prioritise areas that are Emerging or Not Yet Assessed, especially "
            "where surrounding areas are already Secure or Developing."
        )
    elif mode == "steps_only":
        task_text = (
            "Generate 4-6 practical small-step teaching activities.\n"
            "For each provide:\n"
            "- A short activity title\n"
            "- 3-5 small steps (lettered A, B, C…) describing a progression "
            "from initial exposure through to independent demonstration\n\n"
            "Focus on Emerging or Not Yet Assessed statements. "
            "Steps should be concrete classroom actions, not abstract goals."
        )
    else:  # strategies
        task_text = (
            "Suggest 4-6 specific teaching strategies or approaches.\n"
            "For each provide:\n"
            "- A strategy title\n"
            "- A brief description of how to implement it in the classroom\n\n"
            "Focus on areas that need development. Be practical and specific."
        )

    prompt = (
        f"Student: {student.full_name}\n"
        f"Pathway: {student.get_pathway_display()}, Phase {student.phase}\n"
        f"Subject: {subject.name}\n"
        f"Assessment framework(s): {fw_text}\n"
        f"{cohort_text}\n"
        f"Current assessment status:\n{data_text}\n\n"
        f"{task_text}\n\n"
        "IMPORTANT: Respond ONLY with valid JSON — no markdown, no code fences.\n"
        "Use this exact structure:\n"
        '[{"title": "...", "rationale": "...", "small_steps": [{"letter": "A", "description": "..."}]}]\n'
        "If mode is strategies, omit small_steps and just provide title and rationale."
    )

    system = (
        "You are a UK special educational needs (SEN) teaching specialist. "
        "You advise on next steps, targets, and small steps for individual students "
        f"within the {fw_text} framework(s). "
        "Always respond with valid JSON only — no prose, no markdown fencing."
    )

    reply, err = ai_chat(prompt, system=system, max_tokens=2000, timeout=60)
    if err:
        return JsonResponse({"error": err})

    # Parse the JSON response
    reply = reply.strip()
    # Strip markdown code fences if the model wraps them
    if reply.startswith("```"):
        reply = reply.split("\n", 1)[-1]
    if reply.endswith("```"):
        reply = reply.rsplit("```", 1)[0]
    reply = reply.strip()

    try:
        suggestions = json.loads(reply)
    except json.JSONDecodeError:
        return JsonResponse({"error": "AI returned invalid format. Please try again."})

    if not isinstance(suggestions, list):
        suggestions = [suggestions]

    return JsonResponse({"suggestions": suggestions, "mode": mode})


@login_required
@require_POST
def ai_save_suggestions(request, student_id, subject_id):
    """Save AI-generated targets + small steps as real EHCP targets."""
    from evidence.models import EHCPTarget, SmallStep, EHCPOutcome
    from django.utils import timezone

    student = get_object_or_404(Student, pk=student_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    try:
        body = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({"ok": False, "msg": "Invalid request."})

    targets_data = body.get("targets", [])
    outcome_id = body.get("outcome_id")

    outcome = None
    if outcome_id:
        outcome = EHCPOutcome.objects.filter(pk=outcome_id).first()

    created_targets = 0
    created_steps = 0
    today = timezone.now().date()

    for t in targets_data:
        title = (t.get("title") or "").strip()
        if not title:
            continue

        target = EHCPTarget.objects.create(
            student=student,
            outcome=outcome,
            title=title,
            description=f"AI-suggested target for {subject.name}",
            status="NOT_STARTED",
            set_date=today,
            created_by=request.user,
        )

        # Link to relevant assessment areas
        areas = AssessmentArea.objects.filter(subject=subject)
        if areas.exists():
            target.linked_areas.set(areas)

        created_targets += 1

        for step_data in t.get("small_steps", []):
            letter = (step_data.get("letter") or "").strip()
            desc = (step_data.get("description") or "").strip()
            if letter and desc:
                SmallStep.objects.create(
                    target=target,
                    letter=letter,
                    description=desc,
                    order=ord(letter[0]) - 64 if letter else 0,
                )
                created_steps += 1

    msg = f"{created_targets} target(s) and {created_steps} small step(s) saved."
    return JsonResponse({"ok": True, "msg": msg})


@login_required
@require_POST
def record_assessment(request, student_id, statement_id):
    """HTMX endpoint: record/update an assessment for a student+statement."""
    student = get_object_or_404(Student, pk=student_id)
    statement = get_object_or_404(AssessmentStatement, pk=statement_id)
    subject = statement.area.subject

    if not _can_assess_student(request.user, student, subject):
        return HttpResponse("Not authorised to assess this student", status=403)

    status = request.POST.get("status", "NYA")
    notes = request.POST.get("notes", "")

    record = AssessmentRecord.objects.create(
        student=student,
        statement=statement,
        status=status,
        assessed_by=request.user,
        assessed_date=timezone.now().date(),
        notes=notes,
    )

    # Return updated status badge via HTMX
    context = {
        "record": record,
        "statement": statement,
        "student": student,
        "status_choices": AssessmentRecord.STATUS_CHOICES,
    }
    return render(request, "assessments/partials/status_badge.html", context)


@login_required
def bulk_assess(request, class_id, statement_id):
    """Bulk assessment: mark multiple students against one statement."""
    class_group = get_object_or_404(ClassGroup, pk=class_id)
    statement = get_object_or_404(AssessmentStatement, pk=statement_id)
    area = statement.area
    students = class_group.students.filter(is_active=True).order_by(
        "last_name", "first_name"
    )

    can_assess = _can_assess_student(request.user, students.first(), area.subject) if students.exists() else False

    # Build ordered list of all statements in this area for navigation
    all_statements = list(
        AssessmentStatement.objects.filter(area=area)
        .select_related("sub_area")
        .order_by("sub_area__order", "sub_area__name", "order", "pk")
    )
    current_idx = next(
        (i for i, s in enumerate(all_statements) if s.pk == statement.pk), 0
    )
    prev_statement = all_statements[current_idx - 1] if current_idx > 0 else None
    next_statement = (
        all_statements[current_idx + 1]
        if current_idx < len(all_statements) - 1
        else None
    )

    # Get current status for each student
    student_statuses = {}
    for student in students:
        latest = (
            AssessmentRecord.objects.filter(student=student, statement=statement)
            .order_by("-assessed_date", "-created_at")
            .first()
        )
        student_statuses[student.pk] = latest

    if request.method == "POST" and can_assess:
        today = timezone.now().date()
        for student in students:
            new_status = request.POST.get(f"status_{student.pk}")
            if new_status and new_status in dict(AssessmentRecord.STATUS_CHOICES):
                AssessmentRecord.objects.create(
                    student=student,
                    statement=statement,
                    status=new_status,
                    assessed_by=request.user,
                    assessed_date=today,
                )
        messages.success(request, "Bulk assessment saved.")
        # After save, auto-advance to next statement if available
        if next_statement:
            return redirect(
                "assessments:bulk_assess",
                class_id=class_id,
                statement_id=next_statement.pk,
            )
        return redirect("assessments:bulk_assess", class_id=class_id, statement_id=statement_id)

    context = {
        "class_group": class_group,
        "statement": statement,
        "area": area,
        "students": students,
        "student_statuses": student_statuses,
        "can_assess": can_assess,
        "status_choices": AssessmentRecord.STATUS_CHOICES,
        "all_statements": all_statements,
        "current_idx": current_idx,
        "total_statements": len(all_statements),
        "prev_statement": prev_statement,
        "next_statement": next_statement,
    }
    return render(request, "assessments/bulk_assess.html", context)


@login_required
@slt_or_subject_lead_required
def manage_statements(request):
    """List all frameworks and areas — subject leads can manage statements."""
    frameworks = AssessmentFramework.objects.filter(is_active=True).prefetch_related(
        "areas__subject"
    )
    context = {"frameworks": frameworks}
    return render(request, "assessments/manage_statements.html", context)


@login_required
@slt_or_subject_lead_required
def create_framework(request):
    """Create a new assessment framework."""
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()
        if name:
            framework, created = AssessmentFramework.objects.get_or_create(
                name=name, defaults={"description": description}
            )
            if created:
                messages.success(request, f'Framework "{name}" created.')
                return redirect("assessments:framework_detail", framework_id=framework.pk)
            else:
                messages.warning(request, f'Framework "{name}" already exists.')
                return redirect("assessments:framework_detail", framework_id=framework.pk)
        messages.error(request, "Framework name is required.")

    return render(request, "assessments/framework_form.html", {"editing": False})


@login_required
@slt_or_subject_lead_required
def edit_framework(request, framework_id):
    """Edit an existing assessment framework."""
    framework = get_object_or_404(AssessmentFramework, pk=framework_id)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()
        if name:
            framework.name = name
            framework.description = description
            framework.save()
            messages.success(request, f'Framework "{name}" updated.')
            return redirect("assessments:framework_detail", framework_id=framework.pk)
        messages.error(request, "Framework name is required.")

    return render(request, "assessments/framework_form.html", {
        "editing": True,
        "framework": framework,
    })


@login_required
@slt_or_subject_lead_required
@require_POST
def delete_framework(request, framework_id):
    """Delete a framework and all its areas/statements."""
    framework = get_object_or_404(AssessmentFramework, pk=framework_id)
    name = framework.name
    framework.delete()
    messages.success(request, f'Framework "{name}" deleted.')
    return redirect("assessments:framework_hub")


@login_required
@slt_or_subject_lead_required
def create_area(request, framework_id):
    """Create a new assessment area within a framework."""
    framework = get_object_or_404(AssessmentFramework, pk=framework_id)
    subjects = Subject.objects.filter(is_active=True)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        subject_id = request.POST.get("subject")
        new_subject_name = request.POST.get("new_subject_name", "").strip()
        year_group = request.POST.get("year_group", "").strip()
        phase = request.POST.get("phase", "").strip()

        # Resolve subject: existing or create new
        subject = None
        if new_subject_name:
            subject, _ = Subject.objects.get_or_create(
                name=new_subject_name,
                defaults={"applicable_pathways": ["PREP"], "applicable_phases": [1, 2]},
            )
        elif subject_id:
            subject = Subject.objects.filter(pk=subject_id).first()

        if name and subject:
            max_order = framework.areas.aggregate(
                max_order=models.Max("order")
            )["max_order"] or 0
            AssessmentArea.objects.create(
                framework=framework,
                subject=subject,
                name=name,
                year_group=int(year_group) if year_group else None,
                phase=int(phase) if phase else None,
                order=max_order + 1,
            )
            messages.success(request, f'Area "{name}" created.')
            return redirect("assessments:framework_detail", framework_id=framework.pk)
        if not name:
            messages.error(request, "Area name is required.")
        if not subject:
            messages.error(request, "Please select or create a subject.")

    return render(request, "assessments/area_form.html", {
        "editing": False,
        "framework": framework,
        "subjects": subjects,
    })


@login_required
@slt_or_subject_lead_required
def edit_area(request, area_id):
    """Edit an existing assessment area."""
    area = get_object_or_404(AssessmentArea, pk=area_id)
    subjects = Subject.objects.filter(is_active=True)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        subject_id = request.POST.get("subject")
        new_subject_name = request.POST.get("new_subject_name", "").strip()
        year_group = request.POST.get("year_group", "").strip()
        phase = request.POST.get("phase", "").strip()

        subject = None
        if new_subject_name:
            subject, _ = Subject.objects.get_or_create(
                name=new_subject_name,
                defaults={"applicable_pathways": ["PREP"], "applicable_phases": [1, 2]},
            )
        elif subject_id:
            subject = Subject.objects.filter(pk=subject_id).first()

        if name and subject:
            area.name = name
            area.subject = subject
            area.year_group = int(year_group) if year_group else None
            area.phase = int(phase) if phase else None
            area.save()
            messages.success(request, f'Area "{name}" updated.')
            return redirect("assessments:framework_detail", framework_id=area.framework_id)
        if not name:
            messages.error(request, "Area name is required.")
        if not subject:
            messages.error(request, "Please select or create a subject.")

    return render(request, "assessments/area_form.html", {
        "editing": True,
        "area": area,
        "framework": area.framework,
        "subjects": subjects,
    })


@login_required
@slt_or_subject_lead_required
@require_POST
def delete_area(request, area_id):
    """Delete an area and all its statements."""
    area = get_object_or_404(AssessmentArea, pk=area_id)
    fw_id = area.framework_id
    name = area.name
    area.delete()
    messages.success(request, f'Area "{name}" deleted.')
    return redirect("assessments:framework_detail", framework_id=fw_id)


@login_required
@slt_or_subject_lead_required
def manage_sub_areas(request, area_id):
    """Add/edit/delete levels (sub-areas) within an assessment area."""
    area = get_object_or_404(AssessmentArea, pk=area_id)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            name = request.POST.get("name", "").strip()
            if name:
                max_order = area.sub_areas.aggregate(
                    max_order=models.Max("order")
                )["max_order"] or 0
                SubArea.objects.create(area=area, name=name, order=max_order + 1)
                messages.success(request, f'Level "{name}" added.')

        elif action == "delete":
            sub_id = request.POST.get("sub_area_id")
            if sub_id:
                area.sub_areas.filter(pk=sub_id).delete()
                messages.success(request, "Level deleted.")

        elif action == "update":
            sub_id = request.POST.get("sub_area_id")
            name = request.POST.get("name", "").strip()
            if sub_id and name:
                area.sub_areas.filter(pk=sub_id).update(name=name)
                messages.success(request, "Level updated.")

        return redirect("assessments:manage_sub_areas", area_id=area_id)

    sub_areas = area.sub_areas.all()
    context = {"area": area, "sub_areas": sub_areas}
    return render(request, "assessments/manage_sub_areas.html", context)


@login_required
@slt_or_subject_lead_required
def edit_area_statements(request, area_id):
    """Add/edit/reorder statements for an assessment area."""
    area = get_object_or_404(AssessmentArea, pk=area_id)
    sub_areas = area.sub_areas.all()

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            text = request.POST.get("statement_text", "").strip()
            sub_area_id = request.POST.get("sub_area_id")
            sub_area = None
            if sub_area_id:
                sub_area = SubArea.objects.filter(pk=sub_area_id, area=area).first()
            if text:
                qs = area.statements.filter(sub_area=sub_area) if sub_area else area.statements.filter(sub_area__isnull=True)
                max_order = qs.aggregate(
                    max_order=models.Max("order")
                )["max_order"] or 0
                AssessmentStatement.objects.create(
                    area=area, sub_area=sub_area, statement_text=text, order=max_order + 1
                )
                messages.success(request, "Statement added.")

        elif action == "delete":
            stmt_id = request.POST.get("statement_id")
            if stmt_id:
                area.statements.filter(pk=stmt_id).delete()
                messages.success(request, "Statement deleted.")

        elif action == "update":
            stmt_id = request.POST.get("statement_id")
            text = request.POST.get("statement_text", "").strip()
            if stmt_id and text:
                area.statements.filter(pk=stmt_id).update(statement_text=text)
                messages.success(request, "Statement updated.")

        elif action == "move":
            stmt_id = request.POST.get("statement_id")
            target_sub_area_id = request.POST.get("target_sub_area_id", "").strip()
            if stmt_id:
                stmt = area.statements.filter(pk=stmt_id).first()
                if stmt:
                    if target_sub_area_id:
                        target_sub = SubArea.objects.filter(pk=target_sub_area_id, area=area).first()
                        if target_sub:
                            stmt.sub_area = target_sub
                            stmt.save(update_fields=["sub_area"])
                            messages.success(request, f"Statement moved to {target_sub.name}.")
                    else:
                        stmt.sub_area = None
                        stmt.save(update_fields=["sub_area"])
                        messages.success(request, "Statement moved to ungrouped.")

        return redirect("assessments:edit_area_statements", area_id=area_id)

    statements = area.statements.all()
    # Group statements by sub_area for display
    grouped_statements = []
    if sub_areas.exists():
        for sa in sub_areas:
            grouped_statements.append({
                "sub_area": sa,
                "statements": statements.filter(sub_area=sa),
            })
        # Also include any statements not assigned to a sub_area
        ungrouped = statements.filter(sub_area__isnull=True)
        if ungrouped.exists():
            grouped_statements.insert(0, {
                "sub_area": None,
                "statements": ungrouped,
            })
    else:
        grouped_statements.append({
            "sub_area": None,
            "statements": statements,
        })

    context = {
        "area": area,
        "statements": statements,
        "sub_areas": sub_areas,
        "grouped_statements": grouped_statements,
    }
    return render(request, "assessments/edit_area_statements.html", context)


@login_required
@slt_or_subject_lead_required
def import_statements_view(request):
    """Upload a CSV to bulk-import assessment statements."""
    if request.method == "POST":
        csv_file = request.FILES.get("csv_file")
        if not csv_file:
            messages.error(request, "Please select a CSV file.")
            return redirect("assessments:import_statements")

        if not csv_file.name.endswith(".csv"):
            messages.error(request, "File must be a CSV.")
            return redirect("assessments:import_statements")

        decoded = csv_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))

        created = 0
        errors = []
        for i, row in enumerate(reader, start=2):
            try:
                framework_name = row.get("framework", "").strip()
                subject_name = row.get("subject", "").strip()
                area_name = row.get("area", "").strip()
                level_name = row.get("level", "").strip()
                statement_text = row.get("statement", "").strip()
                year_group = row.get("year_group", "").strip()
                phase = row.get("phase", "").strip()

                if not all([framework_name, subject_name, area_name, statement_text]):
                    errors.append(f"Row {i}: Missing required fields.")
                    continue

                framework, _ = AssessmentFramework.objects.get_or_create(
                    name=framework_name
                )
                subject, _ = Subject.objects.get_or_create(name=subject_name)
                area, _ = AssessmentArea.objects.get_or_create(
                    framework=framework,
                    subject=subject,
                    name=area_name,
                    defaults={
                        "year_group": int(year_group) if year_group else None,
                        "phase": int(phase) if phase else None,
                    },
                )
                sub_area = None
                if level_name:
                    sub_area, _ = SubArea.objects.get_or_create(
                        area=area, name=level_name,
                    )
                AssessmentStatement.objects.get_or_create(
                    area=area,
                    sub_area=sub_area,
                    statement_text=statement_text,
                    defaults={"order": created},
                )
                created += 1

            except Exception as e:
                errors.append(f"Row {i}: {e}")

        if created:
            messages.success(request, f"Successfully imported {created} statements.")
        if errors:
            messages.warning(request, f"{len(errors)} rows had errors: {'; '.join(errors[:5])}")

        return redirect("assessments:import_statements")

    context = {
        "expected_columns": "framework, subject, area, level (optional), statement, year_group (optional), phase (optional)"
    }
    return render(request, "assessments/import_statements.html", context)


@login_required
@slt_or_subject_lead_required
def create_snapshot(request):
    """Create a termly snapshot of current assessment status."""
    if request.method == "POST":
        current_term = Term.get_current()
        if not current_term:
            messages.error(request, "No current term configured.")
            return redirect("core:dashboard")

        students = Student.objects.filter(is_active=True)
        areas = AssessmentArea.objects.all()
        created = 0

        for student in students:
            for area in areas:
                statements = area.statements.all()
                total = statements.count()
                if total == 0:
                    continue

                counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
                for statement in statements:
                    latest = (
                        AssessmentRecord.objects.filter(
                            student=student, statement=statement
                        )
                        .order_by("-assessed_date", "-created_at")
                        .first()
                    )
                    if latest:
                        counts[latest.status] = counts.get(latest.status, 0) + 1
                    else:
                        counts["NYA"] += 1

                AssessmentSnapshot.objects.update_or_create(
                    student=student,
                    area=area,
                    term=current_term,
                    defaults={
                        "total_statements": total,
                        "secure_count": counts["SEC"],
                        "developing_count": counts["DEV"],
                        "emerging_count": counts["EME"],
                        "not_assessed_count": counts["NYA"],
                        "snapshot_date": timezone.now().date(),
                    },
                )
                created += 1

        messages.success(request, f"Snapshot created: {created} area records captured.")
        return redirect("core:dashboard")

    return render(request, "assessments/create_snapshot.html")


# ── Frameworks Hub ──────────────────────────────────────────────────

@login_required
def framework_hub(request):
    """Central frameworks landing page — browse, search, create."""
    from evidence.models import MAPPConfig

    # Ensure MAPP appears as a framework if a MAPP config exists
    mapp_config = MAPPConfig.get_active()
    if mapp_config:
        AssessmentFramework.objects.get_or_create(
            name=MAPP_FRAMEWORK_NAME,
            defaults={
                "description": (
                    "Mapping and Assessing Personal Progress — tracks "
                    "learning priorities across customisable dimensions "
                    f"({mapp_config.name}, scale {mapp_config.scale_min}–{mapp_config.scale_max})."
                ),
            },
        )

    q = request.GET.get("q", "").strip()
    frameworks = AssessmentFramework.objects.filter(is_active=True).prefetch_related(
        "areas__statements", "assignments__class_group", "assignments__student",
    )
    if q:
        frameworks = frameworks.filter(name__icontains=q)

    # Annotate with counts
    framework_data = []
    for fw in frameworks:
        is_mapp = fw.name == MAPP_FRAMEWORK_NAME
        areas = fw.areas.all()
        stmt_count = sum(a.statements.count() for a in areas)
        subjects = sorted({a.subject.name for a in areas})
        class_assignments = fw.assignments.filter(class_group__isnull=False).select_related("class_group")
        student_assignments = fw.assignments.filter(student__isnull=False).select_related("student")

        item = {
            "framework": fw,
            "area_count": len(areas),
            "statement_count": stmt_count,
            "subjects": subjects,
            "class_count": class_assignments.count(),
            "student_count": student_assignments.count(),
            "is_mapp": is_mapp,
        }
        if is_mapp and mapp_config:
            dims = list(mapp_config.dimension_configs.values_list("name", flat=True))
            item["mapp_dimensions"] = dims
            item["mapp_scale"] = f"{mapp_config.scale_min}–{mapp_config.scale_max}"

        framework_data.append(item)

    context = {
        "framework_data": framework_data,
        "q": q,
    }
    return render(request, "assessments/framework_hub.html", context)


@login_required
def framework_detail(request, framework_id):
    """Framework detail / builder page — areas, statements, and assignments."""
    from evidence.models import MAPPConfig

    framework = get_object_or_404(
        AssessmentFramework.objects.prefetch_related(
            "areas__subject", "areas__statements",
        ),
        pk=framework_id,
    )
    is_mapp = framework.name == MAPP_FRAMEWORK_NAME
    areas = framework.areas.all()
    class_assignments = (
        FrameworkAssignment.objects.filter(
            framework=framework, class_group__isnull=False
        ).select_related("class_group", "assigned_by")
    )
    student_assignments = (
        FrameworkAssignment.objects.filter(
            framework=framework, student__isnull=False
        ).select_related("student", "assigned_by")
    )
    class_groups = ClassGroup.objects.order_by("year_group", "name")
    students = Student.objects.filter(is_active=True).order_by(
        "class_group__name", "last_name", "first_name"
    )
    subjects = Subject.objects.filter(is_active=True)

    # IDs already assigned so they can be excluded from checkboxes
    assigned_class_ids = set(class_assignments.values_list("class_group_id", flat=True))
    assigned_student_ids = set(student_assignments.values_list("student_id", flat=True))

    context = {
        "framework": framework,
        "areas": areas,
        "class_assignments": class_assignments,
        "student_assignments": student_assignments,
        "class_groups": class_groups,
        "students": students,
        "subjects": subjects,
        "assigned_class_ids": assigned_class_ids,
        "assigned_student_ids": assigned_student_ids,
        "is_mapp": is_mapp,
    }

    if is_mapp:
        mapp_config = MAPPConfig.get_active()
        if mapp_config:
            context["mapp_config"] = mapp_config
            context["mapp_dimensions"] = mapp_config.dimension_configs.all()

    return render(request, "assessments/framework_detail.html", context)


@login_required
@slt_or_subject_lead_required
def mapp_configure(request):
    """In-app page to configure the active MAPP config and its dimensions."""
    from evidence.models import MAPPConfig, MAPPDimensionConfig

    mapp_config = MAPPConfig.get_active()
    if not mapp_config:
        messages.error(request, "No active MAPP configuration found.")
        return redirect("assessments:framework_hub")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "update_config":
            name = request.POST.get("name", "").strip()
            scale_min = request.POST.get("scale_min")
            scale_max = request.POST.get("scale_max")
            if name and scale_min and scale_max:
                mapp_config.name = name
                mapp_config.scale_min = int(scale_min)
                mapp_config.scale_max = int(scale_max)
                mapp_config.save()
                messages.success(request, "MAPP configuration updated.")
            else:
                messages.error(request, "All fields are required.")

        elif action == "add_dimension":
            code = request.POST.get("code", "").strip().upper()
            dim_name = request.POST.get("dim_name", "").strip()
            order = request.POST.get("order", "0")
            if code and dim_name:
                if mapp_config.dimension_configs.filter(code=code).exists():
                    messages.error(request, f"Dimension code '{code}' already exists.")
                else:
                    MAPPDimensionConfig.objects.create(
                        config=mapp_config, code=code, name=dim_name,
                        order=int(order) if order.isdigit() else 0,
                    )
                    messages.success(request, f"Dimension '{code}' added.")
            else:
                messages.error(request, "Code and name are required.")

        elif action == "update_dimension":
            dim_id = request.POST.get("dimension_id")
            dim = MAPPDimensionConfig.objects.filter(pk=dim_id, config=mapp_config).first()
            if dim:
                code = request.POST.get("code", "").strip().upper()
                dim_name = request.POST.get("dim_name", "").strip()
                order = request.POST.get("order", "0")
                if code and dim_name:
                    dim.code = code
                    dim.name = dim_name
                    dim.order = int(order) if order.isdigit() else 0
                    dim.save()
                    messages.success(request, f"Dimension '{code}' updated.")

        elif action == "delete_dimension":
            dim_id = request.POST.get("dimension_id")
            dim = MAPPDimensionConfig.objects.filter(pk=dim_id, config=mapp_config).first()
            if dim:
                dim.delete()
                messages.success(request, "Dimension removed.")

        return redirect("assessments:mapp_configure")

    dimensions = mapp_config.dimension_configs.order_by("order")
    # Find the MAPP framework record for the breadcrumb link
    mapp_framework = AssessmentFramework.objects.filter(name=MAPP_FRAMEWORK_NAME).first()

    return render(request, "assessments/mapp_configure.html", {
        "mapp_config": mapp_config,
        "dimensions": dimensions,
        "mapp_framework": mapp_framework,
    })


@login_required
@slt_or_subject_lead_required
@require_POST
def framework_assign(request, framework_id):
    """Assign a framework to class groups and/or students."""
    framework = get_object_or_404(AssessmentFramework, pk=framework_id)

    class_ids = request.POST.getlist("class_ids")
    student_ids = request.POST.getlist("student_ids")
    created = 0

    for cid in class_ids:
        cg = ClassGroup.objects.filter(pk=cid).first()
        if cg:
            _, was_created = FrameworkAssignment.objects.get_or_create(
                framework=framework, class_group=cg,
                defaults={"assigned_by": request.user},
            )
            if was_created:
                created += 1

    for sid in student_ids:
        st = Student.objects.filter(pk=sid, is_active=True).first()
        if st:
            _, was_created = FrameworkAssignment.objects.get_or_create(
                framework=framework, student=st,
                defaults={"assigned_by": request.user},
            )
            if was_created:
                created += 1

    if created:
        messages.success(request, f"Framework assigned ({created} new assignment{'s' if created != 1 else ''}).")
    else:
        messages.info(request, "No new assignments — already assigned.")
    return redirect("assessments:framework_detail", framework_id=framework_id)


@login_required
@slt_or_subject_lead_required
@require_POST
def framework_unassign(request, assignment_id):
    """Remove a single framework assignment."""
    assignment = get_object_or_404(FrameworkAssignment, pk=assignment_id)
    fw_id = assignment.framework_id
    assignment.delete()
    messages.success(request, "Assignment removed.")
    return redirect("assessments:framework_detail", framework_id=fw_id)


@login_required
@slt_or_subject_lead_required
def personalise_framework(request, framework_id, student_id):
    """Select which statements from a framework are relevant for a student."""
    framework = get_object_or_404(
        AssessmentFramework.objects.prefetch_related("areas__subject", "areas__statements"),
        pk=framework_id,
    )
    student = get_object_or_404(Student, pk=student_id)
    areas = framework.areas.all()

    pf, _ = PersonalisedFramework.objects.get_or_create(
        student=student, framework=framework,
    )
    selected_ids = set(pf.statements.values_list("pk", flat=True))

    if request.method == "POST":
        chosen_ids = set(map(int, request.POST.getlist("statement_ids")))
        # Validate that all chosen IDs belong to this framework
        valid_ids = set(
            AssessmentStatement.objects.filter(
                area__framework=framework, pk__in=chosen_ids
            ).values_list("pk", flat=True)
        )
        pf.statements.set(valid_ids)
        messages.success(
            request,
            f"Personalisation saved — {len(valid_ids)} statement{'s' if len(valid_ids) != 1 else ''} selected for {student.full_name}.",
        )
        return redirect(
            "assessments:personalise_framework",
            framework_id=framework_id,
            student_id=student_id,
        )

    # Count total statements for the summary
    total_statements = sum(a.statements.count() for a in areas)

    return render(request, "assessments/personalise_framework.html", {
        "framework": framework,
        "student": student,
        "areas": areas,
        "selected_ids": selected_ids,
        "total_statements": total_statements,
    })


# ── Class Progress ──────────────────────────────────────────────────


def _cp_date_range(request):
    """Parse date-range filters for class progress. Returns (date_from, date_to, period_label)."""
    from datetime import date as dt_date

    raw_from = request.GET.get("date_from", "").strip()
    raw_to = request.GET.get("date_to", "").strip()

    if raw_from or raw_to:
        df = dt_date.fromisoformat(raw_from) if raw_from else None
        dt = dt_date.fromisoformat(raw_to) if raw_to else None
        label = ""
        if df and dt:
            label = f"{df:%d-%m-%Y} – {dt:%d-%m-%Y}"
        elif df:
            label = f"From {df:%d-%m-%Y}"
        elif dt:
            label = f"Up to {dt:%d-%m-%Y}"
        return df, dt, label

    term_id = request.GET.get("term", "").strip()
    if term_id:
        try:
            term = Term.objects.select_related("academic_year").get(pk=int(term_id))
            return term.start_date, term.end_date, str(term)
        except (Term.DoesNotExist, ValueError):
            pass

    ay_id = request.GET.get("academic_year", "").strip()
    if ay_id:
        try:
            ay = AcademicYear.objects.get(pk=int(ay_id))
            return ay.start_date, ay.end_date, ay.name
        except (AcademicYear.DoesNotExist, ValueError):
            pass

    # Default: all time (no date filtering)
    return None, None, "All time"


def _cp_build_data(request, class_group, date_from=None, date_to=None):
    """Core data builder for class progress — used by the page and exports.

    Returns dict with keys: students, area_data, subjects, selected_subject,
    selected_area, chart_data, pathway_choices, phase_choices.
    """
    from django.db.models import OuterRef, Subquery

    students_qs = class_group.students.filter(is_active=True)

    # Pathway filter
    pathways = request.GET.getlist("pathway")
    if pathways:
        students_qs = students_qs.filter(pathway__in=pathways)

    # Phase filter
    phases = request.GET.getlist("phase")
    if phases:
        students_qs = students_qs.filter(phase__in=[int(p) for p in phases])

    students = list(students_qs.order_by("last_name", "first_name"))
    student_ids = [s.pk for s in students]

    # Frameworks assigned to this class
    assigned_fw_ids = set(
        FrameworkAssignment.objects.filter(
            class_group=class_group, framework__is_active=True
        ).values_list("framework_id", flat=True)
    )

    if assigned_fw_ids:
        areas = AssessmentArea.objects.filter(
            framework_id__in=assigned_fw_ids
        ).select_related("subject", "framework").prefetch_related("statements")
    else:
        areas = AssessmentArea.objects.filter(
            framework__is_active=True
        ).select_related("subject", "framework").prefetch_related("statements")

    # Subject filter
    subject_id = request.GET.get("subject")
    subjects = Subject.objects.filter(
        pk__in=areas.values_list("subject_id", flat=True).distinct()
    ).order_by("name")
    if subject_id:
        areas = areas.filter(subject_id=subject_id)

    # Area filter
    selected_area = request.GET.get("area")
    all_areas_for_picker = list(areas)
    if selected_area:
        areas = areas.filter(pk=selected_area)

    # Build latest-status lookup within the date window
    rec_qs = AssessmentRecord.objects.filter(student_id__in=student_ids)
    if date_from:
        rec_qs = rec_qs.filter(assessed_date__gte=date_from)
    if date_to:
        rec_qs = rec_qs.filter(assessed_date__lte=date_to)

    # Infer per-student, per-subject working level from highest non-NYA level.
    student_subject_floor = {
        (row["student_id"], row["statement__area__subject_id"]): row["max_order"]
        for row in (
            rec_qs.filter(
                statement__sub_area__isnull=False,
                status__in=("EME", "DEV", "SEC"),
            )
            .values("student_id", "statement__area__subject_id")
            .annotate(max_order=models.Max("statement__sub_area__order"))
        )
        if row["max_order"] is not None
    }

    student_area_floor = {
        (row["student_id"], row["statement__area_id"]): row["max_order"]
        for row in (
            rec_qs.filter(
                statement__sub_area__isnull=False,
                status__in=("EME", "DEV", "SEC"),
            )
            .values("student_id", "statement__area_id")
            .annotate(max_order=models.Max("statement__sub_area__order"))
        )
        if row["max_order"] is not None
    }

    area_level_labels = {
        (area_id, order): name
        for area_id, order, name in SubArea.objects.filter(area__in=areas)
        .values_list("area_id", "order", "name")
    }

    statement_meta = {
        sid: (subject_id, area_id, sub_order)
        for sid, subject_id, area_id, sub_order in AssessmentStatement.objects.filter(
            area__in=areas
        ).values_list("pk", "area__subject_id", "area_id", "sub_area__order")
    }

    def _statement_visible_for_student(student_id, statement_id):
        subject_id, _area_id, sub_order = statement_meta.get(
            statement_id, (None, None, None)
        )
        floor = student_subject_floor.get((student_id, subject_id))
        if floor is None:
            return True
        return sub_order is not None and sub_order == floor

    def _build_row_payload(student, level_label, counts, total, extra_levels=None, collapse_id=""):
        return {
            "student": student,
            "level_label": level_label,
            "counts": counts,
            "total": total,
            "pct_secure": round(counts["SEC"] / total * 100) if total else 0,
            "pct_developing": round(counts["DEV"] / total * 100) if total else 0,
            "pct_emerging": round(counts["EME"] / total * 100) if total else 0,
            "pct_nya": round(counts["NYA"] / total * 100) if total else 0,
            "extra_levels": extra_levels or [],
            "collapse_id": collapse_id,
        }

    latest_records = (
        rec_qs.values("student_id", "statement_id")
        .annotate(latest_id=models.Max("id"))
    )
    latest_ids = [r["latest_id"] for r in latest_records]
    status_lookup = {}
    if latest_ids:
        for rec in AssessmentRecord.objects.filter(id__in=latest_ids).values(
            "student_id", "statement_id", "status"
        ):
            status_lookup[(rec["student_id"], rec["statement_id"])] = rec["status"]

    # Build per-student, per-area summary
    area_data = []
    for area in areas:
        stmt_ids = [s.pk for s in area.statements.all()]
        if not stmt_ids:
            continue
        has_levels = SubArea.objects.filter(area=area).exists()
        area_level_orders = list(
            SubArea.objects.filter(area=area)
            .order_by("order", "name")
            .values_list("order", flat=True)
        )
        level_stmt_ids = {}
        for sid in stmt_ids:
            _subject_id, area_id, sub_order = statement_meta.get(sid, (None, None, None))
            if area_id != area.pk or sub_order is None:
                continue
            level_stmt_ids.setdefault(sub_order, []).append(sid)

        student_rows = []
        area_totals = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0, "total": 0}
        for student in students:
            if has_levels:
                per_level_rows = []
                for order in area_level_orders:
                    sids = level_stmt_ids.get(order, [])
                    if not sids:
                        continue
                    level_counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
                    for sid in sids:
                        st = status_lookup.get((student.pk, sid), "NYA")
                        level_counts[st] += 1

                    # For level-based areas, suppress raw NYA volume in class tables.
                    level_counts["NYA"] = 0
                    level_total = level_counts["SEC"] + level_counts["DEV"] + level_counts["EME"]

                    if level_total > 0:
                        per_level_rows.append({
                            "order": order,
                            "level_label": area_level_labels.get((area.pk, order)),
                            "counts": level_counts,
                            "total": level_total,
                            "pct_secure": round(level_counts["SEC"] / level_total * 100) if level_total else 0,
                            "pct_developing": round(level_counts["DEV"] / level_total * 100) if level_total else 0,
                            "pct_emerging": round(level_counts["EME"] / level_total * 100) if level_total else 0,
                            "pct_nya": 0,
                        })

                if per_level_rows:
                    area_floor = student_area_floor.get((student.pk, area.pk))
                    if area_floor is None:
                        area_floor = max(r["order"] for r in per_level_rows)

                    primary = next(
                        (r for r in per_level_rows if r["order"] == area_floor),
                        max(per_level_rows, key=lambda r: r["order"]),
                    )
                    extras = [r for r in per_level_rows if r["order"] != primary["order"]]
                    extras.sort(key=lambda r: r["order"], reverse=True)

                    student_rows.append(
                        _build_row_payload(
                            student=student,
                            level_label=primary["level_label"],
                            counts=primary["counts"],
                            total=primary["total"],
                            extra_levels=extras,
                            collapse_id=f"area-{area.pk}-student-{student.pk}-levels",
                        )
                    )

                    for k in primary["counts"]:
                        area_totals[k] += primary["counts"][k]
                    area_totals["total"] += primary["total"]
                else:
                    counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
                    student_rows.append(
                        _build_row_payload(
                            student=student,
                            level_label=None,
                            counts=counts,
                            total=0,
                        )
                    )
            else:
                counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
                visible_stmt_ids = [
                    sid for sid in stmt_ids
                    if _statement_visible_for_student(student.pk, sid)
                ]
                for sid in visible_stmt_ids:
                    st = status_lookup.get((student.pk, sid), "NYA")
                    counts[st] += 1
                total = len(visible_stmt_ids)

                student_rows.append(
                    _build_row_payload(
                        student=student,
                        level_label=None,
                        counts=counts,
                        total=total,
                    )
                )
                for k in counts:
                    area_totals[k] += counts[k]
                area_totals["total"] += total

        at = area_totals["total"] or 1
        area_data.append({
            "area": area,
            "student_rows": student_rows,
            "totals": area_totals,
            "pcts": {k: round(v / at * 100) for k, v in area_totals.items() if k != "total"},
        })

    # Chart data
    chart_qs = AssessmentRecord.objects.filter(
        student_id__in=student_ids
    ).order_by("assessed_date")
    if date_from:
        chart_qs = chart_qs.filter(assessed_date__gte=date_from)
    if date_to:
        chart_qs = chart_qs.filter(assessed_date__lte=date_to)
    if subject_id:
        chart_qs = chart_qs.filter(statement__area__subject_id=subject_id)
    elif assigned_fw_ids:
        chart_qs = chart_qs.filter(statement__area__framework_id__in=assigned_fw_ids)
    if selected_area:
        chart_qs = chart_qs.filter(statement__area_id=selected_area)

    student_name_map = {s.pk: s.full_name for s in students}
    chart_raw = {}
    for rec in chart_qs.values("student_id", "assessed_date", "status"):
        name = student_name_map.get(rec["student_id"])
        if not name:
            continue
        d = rec["assessed_date"].isoformat()
        chart_raw.setdefault(name, {}).setdefault(d, []).append(
            STATUS_VALUE.get(rec["status"], 0)
        )
    chart_data = {}
    for name, date_map in chart_raw.items():
        sorted_dates = sorted(date_map.keys())
        cum_total = 0
        cum_count = 0
        cum_counts = {"sec": 0, "dev": 0, "eme": 0, "nya": 0}
        val_to_key = {3: "sec", 2: "dev", 1: "eme", 0: "nya"}
        points = []
        for d in sorted_dates:
            day_count = len(date_map[d])
            for v in date_map[d]:
                cum_total += v
                cum_count += 1
                cum_counts[val_to_key[v]] += 1
            points.append({
                "date": d,
                "avg": round(cum_total / cum_count, 2) if cum_count else 0,
                "count": cum_count,
                "day": day_count,
                "sec": cum_counts["sec"],
                "dev": cum_counts["dev"],
                "eme": cum_counts["eme"],
                "nya": cum_counts["nya"],
            })
        chart_data[name] = points

    return {
        "students": students,
        "area_data": area_data,
        "all_areas": all_areas_for_picker,
        "subjects": subjects,
        "selected_subject": subject_id,
        "selected_area": selected_area,
        "chart_data": chart_data,
        "chart_data_json": json.dumps(chart_data),
        "pathway_choices": Student.PATHWAY_CHOICES,
        "phase_choices": Student.PHASE_CHOICES,
        "selected_pathways": pathways,
        "selected_phases": phases,
    }


def _cp_chart_context(request, class_group, data):
    """Build the chart-partial context (subject/area/level breakdowns + trend) for a class.

    Mirrors students.views.student_progress so the partial templates/students/_progress_charts.html
    can render for an entire class with one line per student on the trend chart.
    """
    students = data["students"]
    student_ids = [s.pk for s in students]

    assigned_fw_ids = set(
        FrameworkAssignment.objects.filter(
            class_group=class_group, framework__is_active=True
        ).values_list("framework_id", flat=True)
    )

    # Subjects available for this class
    if assigned_fw_ids:
        subjects = list(
            Subject.objects.filter(
                pk__in=AssessmentArea.objects.filter(
                    framework_id__in=assigned_fw_ids
                ).values_list("subject_id", flat=True).distinct()
            ).order_by("name")
        )
    else:
        subjects = list(data["subjects"])

    # Resolve focus
    focus_subject = None
    focus_area = None
    subj_id = request.GET.get("subject")
    if subj_id:
        try:
            focus_subject = Subject.objects.get(pk=int(subj_id))
        except (Subject.DoesNotExist, ValueError):
            focus_subject = None
    area_id = request.GET.get("area")
    if area_id and focus_subject:
        try:
            focus_area = AssessmentArea.objects.get(pk=int(area_id), subject=focus_subject)
        except (AssessmentArea.DoesNotExist, ValueError):
            focus_area = None

    # Infer per-student, per-subject working level for class charts/tables.
    floor_qs = AssessmentRecord.objects.filter(student_id__in=student_ids)
    student_subject_floor = {
        (row["student_id"], row["statement__area__subject_id"]): row["max_order"]
        for row in (
            floor_qs.filter(
                statement__sub_area__isnull=False,
                status__in=("EME", "DEV", "SEC"),
            )
            .values("student_id", "statement__area__subject_id")
            .annotate(max_order=models.Max("statement__sub_area__order"))
        )
        if row["max_order"] is not None
    }

    statement_meta = {
        sid: (subject_id, sub_order)
        for sid, subject_id, sub_order in AssessmentStatement.objects.filter(
            area__framework_id__in=assigned_fw_ids if assigned_fw_ids else AssessmentArea.objects.filter(framework__is_active=True).values_list("framework_id", flat=True)
        ).values_list("pk", "area__subject_id", "sub_area__order")
    }

    def _statement_visible_for_student(student_id, statement_id):
        subject_id, sub_order = statement_meta.get(statement_id, (None, None))
        floor = student_subject_floor.get((student_id, subject_id))
        if floor is None:
            return True
        return sub_order is not None and sub_order == floor

    def _latest_map(statement_ids):
        ids = list(statement_ids)
        if not student_ids or not ids:
            return {}
        latest = {}
        for r in (
            AssessmentRecord.objects
            .filter(student_id__in=student_ids, statement_id__in=ids)
            .order_by("-assessed_date", "-id")
            .values("student_id", "statement_id", "status")
        ):
            k = (r["student_id"], r["statement_id"])
            if k not in latest:
                latest[k] = r["status"]
        return latest

    def _counts_for(statement_ids):
        ids = list(statement_ids)
        latest = _latest_map(ids)
        counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
        total = 0
        for sid in ids:
            for stu_id in student_ids:
                if not _statement_visible_for_student(stu_id, sid):
                    continue
                counts[latest.get((stu_id, sid), "NYA")] += 1
                total += 1
        return counts, total

    def _area_qs(subject):
        qs = AssessmentArea.objects.filter(subject=subject)
        if assigned_fw_ids:
            qs = qs.filter(framework_id__in=assigned_fw_ids)
        return qs.order_by("name")

    # Per-subject summary
    subject_summaries = []
    for subj in subjects:
        stmt_qs = AssessmentStatement.objects.filter(area__subject=subj)
        if assigned_fw_ids:
            stmt_qs = stmt_qs.filter(area__framework_id__in=assigned_fw_ids)
        counts, total = _counts_for(stmt_qs.values_list("pk", flat=True))
        subject_summaries.append({
            "subject": subj, "total": total,
            "secure": counts["SEC"], "developing": counts["DEV"],
            "emerging": counts["EME"], "not_yet": counts["NYA"],
        })

    # Area breakdown
    area_breakdown = []
    if focus_subject:
        for area in _area_qs(focus_subject):
            counts, total = _counts_for(area.statements.values_list("pk", flat=True))
            area_breakdown.append({
                "area": area, "total": total,
                "secure": counts["SEC"], "developing": counts["DEV"],
                "emerging": counts["EME"], "not_yet": counts["NYA"],
            })

    # Level breakdown (focus area)
    level_breakdown = []
    if focus_area:
        sub_areas = list(SubArea.objects.filter(area=focus_area).order_by("name"))
        groups = [(sa.name, list(sa.statements.values_list("pk", flat=True))) for sa in sub_areas]
        unsorted_ids = list(
            focus_area.statements.filter(sub_area__isnull=True).values_list("pk", flat=True)
        )
        if unsorted_ids:
            groups.append(("(no level)", unsorted_ids))
        for name, ids in groups:
            counts, total = _counts_for(ids)
            if total == 0:
                continue
            level_breakdown.append({
                "name": name, "total": total,
                "secure": counts["SEC"], "developing": counts["DEV"],
                "emerging": counts["EME"], "not_yet": counts["NYA"],
            })

    # Topic + level breakdown (focus subject, no focus area)
    topic_level_breakdown = []
    if focus_subject and not focus_area:
        for area in _area_qs(focus_subject):
            sub_areas = list(SubArea.objects.filter(area=area).order_by("name"))
            for sa in sub_areas:
                counts, total = _counts_for(sa.statements.values_list("pk", flat=True))
                if total == 0:
                    continue
                topic_level_breakdown.append({
                    "name": f"{area.name} — {sa.name}",
                    "total": total,
                    "secure": counts["SEC"], "developing": counts["DEV"],
                    "emerging": counts["EME"], "not_yet": counts["NYA"],
                })

    # Overall totals, focus-aware
    if focus_area:
        src = level_breakdown
    elif focus_subject:
        src = area_breakdown
    else:
        src = subject_summaries
    overall_total = sum(x["total"] for x in src)
    overall_secure = sum(x["secure"] for x in src)
    overall_developing = sum(x["developing"] for x in src)
    overall_emerging = sum(x["emerging"] for x in src)
    overall_nya = sum(x["not_yet"] for x in src)

    # Trend — one series per student (% Secure by term)
    snap_qs = AssessmentSnapshot.objects.filter(student_id__in=student_ids)
    if focus_area:
        snap_qs = snap_qs.filter(area=focus_area)
    elif focus_subject:
        snap_qs = snap_qs.filter(area__subject=focus_subject)
    snap_qs = snap_qs.select_related("term__academic_year", "student").order_by(
        "term__academic_year__start_date", "term__start_date"
    )
    term_order = []
    trend_map = {}
    for snap in snap_qs:
        term_label = str(snap.term)
        if term_label not in term_order:
            term_order.append(term_label)
        name = snap.student.full_name
        bucket = trend_map.setdefault(name, {})
        agg = bucket.setdefault(term_label, {"sec": 0, "total": 0})
        agg["sec"] += snap.secure_count
        agg["total"] += snap.total_statements
    trend_data = {
        name: [
            round((bucket[t]["sec"] / bucket[t]["total"]) * 100, 1)
            if t in bucket and bucket[t]["total"] else None
            for t in term_order
        ]
        for name, bucket in trend_map.items()
    }

    # JSON shapes for the partial
    def _to_chart(items, key):
        return [{
            "name": getattr(it[key], "name", str(it[key])),
            "id": getattr(it[key], "pk", None),
            "secure": it["secure"], "developing": it["developing"],
            "emerging": it["emerging"], "not_yet": it["not_yet"], "total": it["total"],
        } for it in items]

    def _to_chart_named(items):
        return [{
            "name": it["name"],
            "secure": it["secure"], "developing": it["developing"],
            "emerging": it["emerging"], "not_yet": it["not_yet"], "total": it["total"],
        } for it in items]

    return {
        # Override the subjects queryset from _cp_build_data with the focus-friendly list
        "subjects": subjects,
        "subject_summaries": subject_summaries,
        "subject_permissions": {},
        "focus_subject": focus_subject,
        "focus_area": focus_area,
        "area_breakdown": area_breakdown,
        "level_breakdown": level_breakdown,
        "topic_level_breakdown": topic_level_breakdown,
        "subject_chart_data": _to_chart(subject_summaries, "subject"),
        "area_chart_data": _to_chart(area_breakdown, "area") if area_breakdown else [],
        "level_chart_data": _to_chart_named(level_breakdown),
        "topic_level_chart_data": _to_chart_named(topic_level_breakdown),
        "trend_terms": term_order,
        "trend_data": trend_data,
        "overall_total": overall_total,
        "overall_secure": overall_secure,
        "overall_developing": overall_developing,
        "overall_emerging": overall_emerging,
        "overall_nya": overall_nya,
    }


@login_required
def class_progress(request, class_id):
    """Show assessment progress for all students in a class, broken down by subject/area."""
    class_group = get_object_or_404(ClassGroup, pk=class_id)
    date_from, date_to, period_label = _cp_date_range(request)
    data = _cp_build_data(request, class_group, date_from, date_to)

    # View toggle
    view_mode = request.GET.get("view", "both")  # "chart", "table", "both"

    # Comparison mode — load a second period
    compare = request.GET.get("compare")  # term PK to compare against
    compare_data = None
    compare_label = ""
    if compare:
        try:
            cmp_term = Term.objects.select_related("academic_year").get(pk=int(compare))
            compare_data = _cp_build_data(request, class_group, cmp_term.start_date, cmp_term.end_date)
            compare_label = str(cmp_term)
        except (Term.DoesNotExist, ValueError):
            pass

    chart_ctx = _cp_chart_context(request, class_group, data)

    context = {
        "class_group": class_group,
        "view_mode": view_mode,
        "period_label": period_label,
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
        "academic_years": AcademicYear.objects.all(),
        "terms": Term.objects.select_related("academic_year").all(),
        "selected_academic_year": request.GET.get("academic_year", ""),
        "selected_term": request.GET.get("term", ""),
        "compare": compare,
        "compare_data": compare_data,
        "compare_label": compare_label,
        **data,
        **chart_ctx,
    }

    if request.GET.get("partial") == "1" or request.headers.get("HX-Request"):
        return render(request, "students/_progress_charts.html", context)
    return render(request, "assessments/class_progress.html", context)


@login_required
def class_progress_export_excel(request, class_id):
    """Export class progress as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return HttpResponse("openpyxl is not installed.", status=500)

    class_group = get_object_or_404(ClassGroup, pk=class_id)
    date_from, date_to, period_label = _cp_date_range(request)
    data = _cp_build_data(request, class_group, date_from, date_to)

    fills = {
        "SEC": PatternFill(start_color="198754", end_color="198754", fill_type="solid"),
        "DEV": PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid"),
        "EME": PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
        "NYA": PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid"),
    }
    white_font = Font(color="FFFFFF", bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Progress"

    # Title rows
    ws.append([f"{class_group.name} — Class Progress"])
    ws.append([f"Period: {period_label}"])
    ws.append([])

    for ad in data["area_data"]:
        ws.append([ad["area"].name, "", "N", "E", "D", "S", "% Secure"])
        ws[ws.max_row][0].font = Font(bold=True, size=12)

        for row in ad["student_rows"]:
            r = [
                row["student"].full_name,
                row["student"].get_pathway_display(),
                row["counts"]["NYA"],
                row["counts"]["EME"],
                row["counts"]["DEV"],
                row["counts"]["SEC"],
                f'{row["pct_secure"]}%',
            ]
            ws.append(r)
            ri = ws.max_row
            status_cols = {"NYA": 3, "EME": 4, "DEV": 5, "SEC": 6}
            for status, col in status_cols.items():
                cell = ws.cell(row=ri, column=col)
                cell.fill = fills[status]
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")

        ws.append([])

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_name = class_group.name.replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="class_progress_{safe_name}.xlsx"'
    wb.save(response)
    return response


@login_required
def class_progress_export_pdf(request, class_id):
    """Export class progress as a PDF."""
    import weasyprint
    from django.template.loader import render_to_string

    class_group = get_object_or_404(ClassGroup, pk=class_id)
    date_from, date_to, period_label = _cp_date_range(request)
    data = _cp_build_data(request, class_group, date_from, date_to)

    context = {
        "class_group": class_group,
        "area_data": data["area_data"],
        "period_label": period_label,
        "today": timezone.now().date(),
    }
    html_string = render_to_string("pdf/class_progress.html", context)
    pdf = weasyprint.HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    safe_name = class_group.name.replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="class_progress_{safe_name}.pdf"'
    return response


# ── Whole-School Progress ───────────────────────────────────────────


@login_required
@slt_or_subject_lead_required
def school_progress(request):
    """Cross-class progress overview for school leaders.

    Shows per-class summary rows with NEDS counts and progress bars,
    drill-down by subject/area, and all the same filters as class progress.
    """

    date_from, date_to, period_label = _cp_date_range(request)

    # ── Student-level filters ──
    students_qs = Student.objects.filter(is_active=True)
    pathways = request.GET.getlist("pathway")
    if pathways:
        students_qs = students_qs.filter(pathway__in=pathways)
    phases = request.GET.getlist("phase")
    if phases:
        students_qs = students_qs.filter(phase__in=[int(p) for p in phases])
    year_groups_filter = request.GET.getlist("year_group")
    if year_groups_filter:
        students_qs = students_qs.filter(year_group__in=[int(y) for y in year_groups_filter])

    # Which classes to show
    class_filter = request.GET.getlist("class_group")
    all_classes = ClassGroup.objects.order_by("year_group", "name")
    if class_filter:
        classes = all_classes.filter(pk__in=[int(c) for c in class_filter])
    else:
        classes = all_classes

    # Subject / area
    subject_id = request.GET.get("subject")
    all_subjects = Subject.objects.filter(is_active=True).order_by("name")

    # All active areas
    areas = AssessmentArea.objects.filter(
        framework__is_active=True
    ).select_related("subject", "framework").prefetch_related("statements")
    if subject_id:
        areas = areas.filter(subject_id=subject_id)
    selected_area = request.GET.get("area")
    all_areas_for_picker = list(areas)
    if selected_area:
        areas = areas.filter(pk=selected_area)

    # Collect statement IDs per area
    area_stmt_map = {}
    for area in areas:
        sids = [s.pk for s in area.statements.all()]
        if sids:
            area_stmt_map[area] = sids
    all_stmt_ids = []
    for sids in area_stmt_map.values():
        all_stmt_ids.extend(sids)

    # Build per-class data
    class_data = []
    chart_raw_all = {}  # class_name → {date → [values]}

    # Year groups available for filter checkboxes
    year_groups_available = sorted(
        Student.objects.filter(is_active=True, year_group__isnull=False)
        .values_list("year_group", flat=True).distinct()
    )

    for cg in classes:
        students = list(
            students_qs.filter(class_group=cg).order_by("last_name", "first_name")
        )
        if not students:
            continue
        student_ids = [s.pk for s in students]

        # Latest-status lookup within date window
        rec_qs = AssessmentRecord.objects.filter(student_id__in=student_ids)
        if date_from:
            rec_qs = rec_qs.filter(assessed_date__gte=date_from)
        if date_to:
            rec_qs = rec_qs.filter(assessed_date__lte=date_to)

        latest_records = (
            rec_qs.values("student_id", "statement_id")
            .annotate(latest_id=models.Max("id"))
        )
        latest_ids = [r["latest_id"] for r in latest_records]
        status_lookup = {}
        if latest_ids:
            for rec in AssessmentRecord.objects.filter(id__in=latest_ids).values(
                "student_id", "statement_id", "status"
            ):
                status_lookup[(rec["student_id"], rec["statement_id"])] = rec["status"]

        # Class-wide NEDS totals (across all filtered areas)
        class_totals = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0, "total": 0}
        area_rows = []
        for area, stmt_ids in area_stmt_map.items():
            area_totals = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0, "total": 0}
            for student in students:
                for sid in stmt_ids:
                    st = status_lookup.get((student.pk, sid), "NYA")
                    area_totals[st] += 1
                area_totals["total"] += len(stmt_ids)

            at = area_totals["total"] or 1
            area_rows.append({
                "area": area,
                "totals": dict(area_totals),
                "pcts": {
                    k: round(v / at * 100)
                    for k, v in area_totals.items() if k != "total"
                },
            })
            for k in ("SEC", "DEV", "EME", "NYA"):
                class_totals[k] += area_totals[k]
            class_totals["total"] += area_totals["total"]

        ct = class_totals["total"] or 1
        class_data.append({
            "class_group": cg,
            "student_count": len(students),
            "totals": class_totals,
            "pcts": {
                k: round(v / ct * 100)
                for k, v in class_totals.items() if k != "total"
            },
            "area_rows": area_rows,
        })

        # Chart data: average status per class per date
        chart_qs = rec_qs
        if subject_id:
            chart_qs = chart_qs.filter(statement__area__subject_id=subject_id)
        if selected_area:
            chart_qs = chart_qs.filter(statement__area_id=selected_area)

        date_vals = {}
        for rec in chart_qs.values("assessed_date", "status"):
            d = rec["assessed_date"].isoformat()
            date_vals.setdefault(d, []).append(
                STATUS_VALUE.get(rec["status"], 0)
            )
        if date_vals:
            sorted_dates = sorted(date_vals.keys())
            cum_total = 0
            cum_count = 0
            cum_counts = {"sec": 0, "dev": 0, "eme": 0, "nya": 0}
            val_to_key = {3: "sec", 2: "dev", 1: "eme", 0: "nya"}
            points = []
            for d in sorted_dates:
                day_count = len(date_vals[d])
                for v in date_vals[d]:
                    cum_total += v
                    cum_count += 1
                    cum_counts[val_to_key[v]] += 1
                points.append({
                    "date": d,
                    "avg": round(cum_total / cum_count, 2) if cum_count else 0,
                    "count": cum_count,
                    "day": day_count,
                    "sec": cum_counts["sec"],
                    "dev": cum_counts["dev"],
                    "eme": cum_counts["eme"],
                    "nya": cum_counts["nya"],
                })
            chart_raw_all[cg.name] = points

    # Grand totals across all classes
    grand = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0, "total": 0}
    for cd in class_data:
        for k in ("SEC", "DEV", "EME", "NYA"):
            grand[k] += cd["totals"][k]
        grand["total"] += cd["totals"]["total"]
    gt = grand["total"] or 1
    grand_pcts = {k: round(v / gt * 100) for k, v in grand.items() if k != "total"}

    # Comparison
    compare = request.GET.get("compare")
    compare_data = None
    compare_label = ""
    if compare:
        try:
            cmp_term = Term.objects.select_related("academic_year").get(pk=int(compare))
            # Quick rebuild with comparison dates — only need class totals
            cmp_class_data = []
            for cg in classes:
                cmp_students = list(
                    students_qs.filter(class_group=cg).order_by("last_name")
                )
                if not cmp_students:
                    continue
                cmp_sids = [s.pk for s in cmp_students]
                cmp_qs = AssessmentRecord.objects.filter(
                    student_id__in=cmp_sids,
                    assessed_date__gte=cmp_term.start_date,
                    assessed_date__lte=cmp_term.end_date,
                )
                cmp_latest = (
                    cmp_qs.values("student_id", "statement_id")
                    .annotate(latest_id=models.Max("id"))
                )
                cmp_lids = [r["latest_id"] for r in cmp_latest]
                cmp_lookup = {}
                if cmp_lids:
                    for rec in AssessmentRecord.objects.filter(id__in=cmp_lids).values(
                        "student_id", "statement_id", "status"
                    ):
                        cmp_lookup[(rec["student_id"], rec["statement_id"])] = rec["status"]

                ct2 = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0, "total": 0}
                for area, stmt_ids in area_stmt_map.items():
                    for student in cmp_students:
                        for sid in stmt_ids:
                            st = cmp_lookup.get((student.pk, sid), "NYA")
                            ct2[st] += 1
                        ct2["total"] += len(stmt_ids)
                ctt = ct2["total"] or 1
                cmp_class_data.append({
                    "class_group": cg,
                    "student_count": len(cmp_students),
                    "totals": ct2,
                    "pcts": {k: round(v / ctt * 100) for k, v in ct2.items() if k != "total"},
                })
            compare_data = cmp_class_data
            compare_label = str(cmp_term)
        except (Term.DoesNotExist, ValueError):
            pass

    view_mode = request.GET.get("view", "both")

    context = {
        "class_data": class_data,
        "grand_totals": grand,
        "grand_pcts": grand_pcts,
        "view_mode": view_mode,
        "period_label": period_label,
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
        "academic_years": AcademicYear.objects.all(),
        "terms": Term.objects.select_related("academic_year").all(),
        "selected_academic_year": request.GET.get("academic_year", ""),
        "selected_term": request.GET.get("term", ""),
        "subjects": all_subjects,
        "selected_subject": subject_id,
        "all_areas": all_areas_for_picker,
        "selected_area": selected_area,
        "all_classes": all_classes,
        "selected_classes": class_filter,
        "pathway_choices": Student.PATHWAY_CHOICES,
        "phase_choices": Student.PHASE_CHOICES,
        "selected_pathways": pathways,
        "selected_phases": phases,
        "year_groups_available": year_groups_available,
        "selected_year_groups": year_groups_filter,
        "chart_data_json": json.dumps(chart_raw_all),
        "compare": compare,
        "compare_data": compare_data,
        "compare_label": compare_label,
    }
    return render(request, "assessments/school_progress.html", context)


@login_required
@slt_or_subject_lead_required
def school_progress_export_excel(request):
    """Export whole-school progress as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return HttpResponse("openpyxl is not installed.", status=500)

    date_from, date_to, period_label = _cp_date_range(request)

    students_qs = Student.objects.filter(is_active=True)
    pathways = request.GET.getlist("pathway")
    if pathways:
        students_qs = students_qs.filter(pathway__in=pathways)
    phases_f = request.GET.getlist("phase")
    if phases_f:
        students_qs = students_qs.filter(phase__in=[int(p) for p in phases_f])
    yg_f = request.GET.getlist("year_group")
    if yg_f:
        students_qs = students_qs.filter(year_group__in=[int(y) for y in yg_f])

    class_filter = request.GET.getlist("class_group")
    classes = ClassGroup.objects.order_by("year_group", "name")
    if class_filter:
        classes = classes.filter(pk__in=[int(c) for c in class_filter])

    subject_id = request.GET.get("subject")
    areas = AssessmentArea.objects.filter(
        framework__is_active=True
    ).select_related("subject").prefetch_related("statements")
    if subject_id:
        areas = areas.filter(subject_id=subject_id)
    selected_area = request.GET.get("area")
    if selected_area:
        areas = areas.filter(pk=selected_area)

    area_stmt_map = {}
    for area in areas:
        sids = [s.pk for s in area.statements.all()]
        if sids:
            area_stmt_map[area] = sids

    fills = {
        "SEC": PatternFill(start_color="198754", end_color="198754", fill_type="solid"),
        "DEV": PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid"),
        "EME": PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
        "NYA": PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid"),
    }
    white_font = Font(color="FFFFFF", bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "School Progress"
    ws.append(["Whole School Progress"])
    ws.append([f"Period: {period_label}"])
    ws.append([])
    ws.append(["Class", "Students", "NYA", "Emerging", "Developing", "Secure", "% Secure"])
    ws[ws.max_row][0].font = Font(bold=True)

    for cg in classes:
        students = list(students_qs.filter(class_group=cg))
        if not students:
            continue
        student_ids = [s.pk for s in students]

        rec_qs = AssessmentRecord.objects.filter(student_id__in=student_ids)
        if date_from:
            rec_qs = rec_qs.filter(assessed_date__gte=date_from)
        if date_to:
            rec_qs = rec_qs.filter(assessed_date__lte=date_to)

        latest_records = rec_qs.values("student_id", "statement_id").annotate(
            latest_id=models.Max("id")
        )
        latest_ids = [r["latest_id"] for r in latest_records]
        status_lookup = {}
        if latest_ids:
            for rec in AssessmentRecord.objects.filter(id__in=latest_ids).values(
                "student_id", "statement_id", "status"
            ):
                status_lookup[(rec["student_id"], rec["statement_id"])] = rec["status"]

        totals = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0, "total": 0}
        for area, stmt_ids in area_stmt_map.items():
            for student in students:
                for sid in stmt_ids:
                    st = status_lookup.get((student.pk, sid), "NYA")
                    totals[st] += 1
                totals["total"] += len(stmt_ids)

        tt = totals["total"] or 1
        ws.append([
            cg.name,
            len(students),
            totals["NYA"],
            totals["EME"],
            totals["DEV"],
            totals["SEC"],
            f'{round(totals["SEC"] / tt * 100)}%',
        ])
        ri = ws.max_row
        for col, status in [(3, "NYA"), (4, "EME"), (5, "DEV"), (6, "SEC")]:
            cell = ws.cell(row=ri, column=col)
            cell.fill = fills[status]
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="school_progress.xlsx"'
    wb.save(response)
    return response


@login_required
@slt_or_subject_lead_required
def school_progress_export_pdf(request):
    """Export whole-school progress as a PDF."""
    import weasyprint, json as _json
    from django.template.loader import render_to_string

    date_from, date_to, period_label = _cp_date_range(request)

    students_qs = Student.objects.filter(is_active=True)
    pathways = request.GET.getlist("pathway")
    if pathways:
        students_qs = students_qs.filter(pathway__in=pathways)
    phases_f = request.GET.getlist("phase")
    if phases_f:
        students_qs = students_qs.filter(phase__in=[int(p) for p in phases_f])

    class_filter = request.GET.getlist("class_group")
    classes = ClassGroup.objects.order_by("year_group", "name")
    if class_filter:
        classes = classes.filter(pk__in=[int(c) for c in class_filter])

    subject_id = request.GET.get("subject")
    areas = AssessmentArea.objects.filter(
        framework__is_active=True
    ).select_related("subject").prefetch_related("statements")
    if subject_id:
        areas = areas.filter(subject_id=subject_id)
    selected_area = request.GET.get("area")
    if selected_area:
        areas = areas.filter(pk=selected_area)

    area_stmt_map = {}
    for area in areas:
        sids = [s.pk for s in area.statements.all()]
        if sids:
            area_stmt_map[area] = sids

    rows = []
    for cg in classes:
        students = list(students_qs.filter(class_group=cg))
        if not students:
            continue
        student_ids = [s.pk for s in students]

        rec_qs = AssessmentRecord.objects.filter(student_id__in=student_ids)
        if date_from:
            rec_qs = rec_qs.filter(assessed_date__gte=date_from)
        if date_to:
            rec_qs = rec_qs.filter(assessed_date__lte=date_to)

        latest = rec_qs.values("student_id", "statement_id").annotate(
            latest_id=models.Max("id")
        )
        lids = [r["latest_id"] for r in latest]
        lookup = {}
        if lids:
            for rec in AssessmentRecord.objects.filter(id__in=lids).values(
                "student_id", "statement_id", "status"
            ):
                lookup[(rec["student_id"], rec["statement_id"])] = rec["status"]

        totals = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0, "total": 0}
        for area, stmt_ids in area_stmt_map.items():
            for student in students:
                for sid in stmt_ids:
                    totals[lookup.get((student.pk, sid), "NYA")] += 1
                totals["total"] += len(stmt_ids)
        tt = totals["total"] or 1
        rows.append({
            "name": cg.name,
            "students": len(students),
            "sec": totals["SEC"],
            "dev": totals["DEV"],
            "eme": totals["EME"],
            "nya": totals["NYA"],
            "pct_sec": round(totals["SEC"] / tt * 100),
        })

    context = {
        "rows": rows,
        "period_label": period_label,
        "today": timezone.now().date(),
    }
    html_string = render_to_string("pdf/school_progress.html", context)
    pdf = weasyprint.HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="school_progress.pdf"'
    return response
