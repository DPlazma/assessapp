from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, Prefetch
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_POST

from students.models import Student, Subject, ClassGroup
from core.models import AcademicYear, Term
from staff.models import ClassAssignment, ClassCover, SubjectLead
from assessments.views import _can_assess_student
from assessments.models import (
    AssessmentFramework,
    AssessmentArea,
    AssessmentStatement,
    AssessmentRecord,
)


def _get_filter_options():
    """Return all available filter values for the checkbox panel."""
    return {
        "pathways": Student.PATHWAY_CHOICES,
        "phases": Student.PHASE_CHOICES,
        "class_groups": ClassGroup.objects.all(),
        "year_groups": sorted(
            Student.objects.filter(year_group__isnull=False)
            .values_list("year_group", flat=True)
            .distinct()
        ),
        "subjects": Subject.objects.filter(is_active=True),
        "frameworks": AssessmentFramework.objects.filter(is_active=True),
        "academic_years": AcademicYear.objects.all(),
        "terms": Term.objects.select_related("academic_year").all(),
    }


def _get_date_range(request, params=None):
    """Determine the date range from GET params (academic_year, term, date_from, date_to).

    Priority: explicit dates > term > academic year.
    Returns (date_from, date_to, active_time_filters dict).
    """
    params = params or request.GET
    active = {}
    date_from = None
    date_to = None

    # Explicit custom dates take highest priority
    raw_from = params.get("date_from", "").strip()
    raw_to = params.get("date_to", "").strip()
    if raw_from or raw_to:
        from datetime import date as dt_date
        try:
            if raw_from:
                date_from = dt_date.fromisoformat(raw_from)
                active["date_from"] = raw_from
            if raw_to:
                date_to = dt_date.fromisoformat(raw_to)
                active["date_to"] = raw_to
        except ValueError:
            pass
        if date_from or date_to:
            return date_from, date_to, active

    # Term overrides academic year
    term_id = params.get("term", "").strip()
    if term_id:
        try:
            term = Term.objects.get(pk=int(term_id))
            active["term"] = term_id
            return term.start_date, term.end_date, active
        except (Term.DoesNotExist, ValueError):
            pass

    # Academic year
    ay_id = params.get("academic_year", "").strip()
    if ay_id:
        try:
            ay = AcademicYear.objects.get(pk=int(ay_id))
            active["academic_year"] = ay_id
            return ay.start_date, ay.end_date, active
        except (AcademicYear.DoesNotExist, ValueError):
            pass

    return None, None, active


def _apply_date_filter_to_records(qs, date_from, date_to):
    """Narrow an AssessmentRecord queryset to a date window."""
    if date_from:
        qs = qs.filter(assessed_date__gte=date_from)
    if date_to:
        qs = qs.filter(assessed_date__lte=date_to)
    return qs


def _time_label(date_from, date_to, time_filters):
    """Human-readable description of the active time window."""
    if "term" in time_filters:
        try:
            term = Term.objects.select_related("academic_year").get(pk=int(time_filters["term"]))
            return str(term)
        except Term.DoesNotExist:
            pass
    if "academic_year" in time_filters:
        try:
            ay = AcademicYear.objects.get(pk=int(time_filters["academic_year"]))
            return str(ay)
        except AcademicYear.DoesNotExist:
            pass
    if date_from and date_to:
        return f"{date_from:%d-%m-%Y} – {date_to:%d-%m-%Y}"
    if date_from:
        return f"From {date_from:%d-%m-%Y}"
    if date_to:
        return f"Up to {date_to:%d-%m-%Y}"
    return "All time"


def _apply_student_filters(request, qs, params=None):
    """Apply checkbox filters from GET params to a Student queryset.

    Returns (filtered_qs, active_filters dict).
    """
    params = params or request.GET
    active = {}

    pathways = params.getlist("pathway")
    if pathways:
        qs = qs.filter(pathway__in=pathways)
        active["pathway"] = pathways

    phases = params.getlist("phase")
    if phases:
        qs = qs.filter(phase__in=[int(p) for p in phases])
        active["phase"] = phases

    class_ids = params.getlist("class_group")
    if class_ids:
        qs = qs.filter(class_group_id__in=[int(c) for c in class_ids])
        active["class_group"] = class_ids

    year_groups = params.getlist("year_group")
    if year_groups:
        qs = qs.filter(year_group__in=[int(y) for y in year_groups])
        active["year_group"] = year_groups

    return qs, active


def _apply_area_filters(request, qs, params=None):
    """Apply subject/framework filters to an AssessmentArea queryset."""
    params = params or request.GET
    subject_ids = params.getlist("subject")
    if subject_ids:
        qs = qs.filter(subject_id__in=[int(s) for s in subject_ids])

    framework_ids = params.getlist("framework")
    if framework_ids:
        qs = qs.filter(framework_id__in=[int(f) for f in framework_ids])

    return qs


def _cohort_effective_params(request):
    """Apply smart defaults for the cohort report when filters are empty or partial."""
    params = request.GET.copy()
    applied = []

    has_time_filters = any(
        params.get(key, "").strip()
        for key in ("date_from", "date_to", "term", "academic_year")
    )
    has_scope_filters = any(
        params.getlist(key)
        for key in ("pathway", "phase", "class_group", "year_group", "subject", "framework")
    )

    if not has_time_filters:
        current_term = Term.get_current()
        if current_term:
            params.setlist("term", [str(current_term.pk)])
            applied.append(f"Current term: {current_term}")

    if not has_scope_filters:
        profile = getattr(request.user, "staffprofile", None)
        defaulted = False
        if profile and profile.is_subject_lead:
            subject_ids = list(
                SubjectLead.objects.filter(user=request.user).values_list("subject_id", flat=True)
            )
            if subject_ids:
                params.setlist("subject", [str(subject_id) for subject_id in subject_ids])
                applied.append("Your subject lead areas")
                defaulted = True

        if profile and not defaulted and profile.is_pathway_lead and profile.lead_pathway:
            params.setlist("pathway", [profile.lead_pathway])
            applied.append(f"Your pathway: {dict(Student.PATHWAY_CHOICES).get(profile.lead_pathway, profile.lead_pathway)}")
            defaulted = True

        if not defaulted:
            today = timezone.now().date()
            class_ids = set(
                ClassAssignment.objects.filter(user=request.user).values_list("class_group_id", flat=True)
            )
            class_ids.update(
                ClassCover.objects.filter(
                    user=request.user,
                    start_date__lte=today,
                    end_date__gte=today,
                ).values_list("class_group_id", flat=True)
            )
            if class_ids:
                params.setlist("class_group", [str(class_id) for class_id in sorted(class_ids)])
                applied.append("Your assigned classes")

    return params, applied


@login_required
def report_index(request):
    """Landing page for reports — links to each report type."""
    return render(request, "reports/index.html")


@login_required
def cohort_report(request):
    """Grid of students × statements showing latest RAG status per student.

    Filters: pathway, phase, class group, year group, subject, framework.
    """
    filter_options = _get_filter_options()
    effective_params, smart_defaults_applied = _cohort_effective_params(request)
    date_from, date_to, time_filters = _get_date_range(request, effective_params)

    students = Student.objects.filter(is_active=True).select_related("class_group")
    students, active_filters = _apply_student_filters(request, students, effective_params)
    active_filters.update(time_filters)
    students = list(students.order_by("last_name", "first_name"))

    # Get assessment areas (filtered by subject/framework if requested)
    areas = AssessmentArea.objects.select_related("subject", "framework").prefetch_related(
        Prefetch("statements", queryset=AssessmentStatement.objects.order_by("order", "pk"))
    )
    areas = list(_apply_area_filters(request, areas, effective_params))

    # Build list of all statements from filtered areas
    statements = []
    for area in areas:
        for stmt in area.statements.all():
            stmt.report_subject_id = area.subject_id
            statements.append(stmt)

    # Build a grid: {student_id: {statement_id: status}}
    grid = {}
    if students and statements:
        stmt_ids = [s.pk for s in statements]
        student_ids = [student.pk for student in students]

        # Subquery: latest record per student+statement (within date window)
        from django.db.models import Max, Subquery, OuterRef

        latest_qs = AssessmentRecord.objects.filter(
            student_id=OuterRef("student_id"),
            statement_id=OuterRef("statement_id"),
        )
        latest_qs = _apply_date_filter_to_records(latest_qs, date_from, date_to)
        latest_pks = latest_qs.order_by("-assessed_date", "-created_at").values("pk")[:1]

        base_records = AssessmentRecord.objects.filter(
            student_id__in=student_ids,
            statement_id__in=stmt_ids,
            pk__in=Subquery(latest_pks),
        )
        records = base_records.values("student_id", "statement_id", "status")

        for r in records:
            grid.setdefault(r["student_id"], {})[r["statement_id"]] = r["status"]

    # Summary counts per student
    student_summaries = {}
    for student in students:
        s_grid = grid.get(student.pk, {})
        counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
        for stmt in statements:
            status = s_grid.get(stmt.pk, "NYA")
            counts[status] = counts.get(status, 0) + 1
        student_summaries[student.pk] = counts

    editable_by_student = {}
    subjects_by_id = {subject.pk: subject for subject in Subject.objects.filter(
        pk__in=areas.values_list("subject_id", flat=True).distinct()
    )}
    for student in students:
        editable_by_student[student.pk] = {}
        for subject_id, subject in subjects_by_id.items():
            editable_by_student[student.pk][subject_id] = _can_assess_student(
                request.user, student, subject
            )

    # Describe the active time window
    time_label = _time_label(date_from, date_to, time_filters)

    context = {
        "students": students,
        "statements": statements,
        "areas": areas,
        "grid": grid,
        "student_summaries": student_summaries,
        "editable_by_student": editable_by_student,
        "filter_options": filter_options,
        "active_filters": active_filters,
        "total_statements": len(statements),
        "time_label": time_label,
        "smart_defaults_applied": smart_defaults_applied,
        "effective_querystring": effective_params.urlencode(),
    }
    return render(request, "reports/cohort_report.html", context)


@login_required
@require_POST
def cohort_update_cell(request):
    """Quick-update one cohort report cell by creating a fresh AssessmentRecord."""
    student = get_object_or_404(Student, pk=request.POST.get("student_id"))
    statement = get_object_or_404(AssessmentStatement, pk=request.POST.get("statement_id"))
    subject = statement.area.subject

    if not _can_assess_student(request.user, student, subject):
        return JsonResponse({"ok": False, "error": "Not authorised"}, status=403)

    status = (request.POST.get("status") or "NYA").strip().upper()
    if status not in dict(AssessmentRecord.STATUS_CHOICES):
        return JsonResponse({"ok": False, "error": "Invalid status"}, status=400)

    AssessmentRecord.objects.create(
        student=student,
        statement=statement,
        status=status,
        assessed_by=request.user,
        assessed_date=timezone.now().date(),
        notes="Updated from cohort report.",
    )

    return JsonResponse({
        "ok": True,
        "status": status,
        "student_id": student.pk,
        "statement_id": statement.pk,
    })


@login_required
def subject_progress(request):
    """Per-subject breakdown: how many students at each status level.

    Shows % of cohort at Secure / Developing / Emerging / NYA per subject.
    """
    filter_options = _get_filter_options()
    date_from, date_to, time_filters = _get_date_range(request)

    students = Student.objects.filter(is_active=True)
    students, active_filters = _apply_student_filters(request, students)
    active_filters.update(time_filters)
    student_ids = list(students.values_list("pk", flat=True))

    # Filter subjects
    subjects = Subject.objects.filter(is_active=True)
    subject_filter = request.GET.getlist("subject")
    if subject_filter:
        subjects = subjects.filter(pk__in=[int(s) for s in subject_filter])

    framework_filter = request.GET.getlist("framework")

    subject_data = []
    for subject in subjects:
        areas = AssessmentArea.objects.filter(subject=subject)
        if framework_filter:
            areas = areas.filter(framework_id__in=[int(f) for f in framework_filter])

        stmt_ids = list(
            AssessmentStatement.objects.filter(area__in=areas).values_list("pk", flat=True)
        )
        total_possible = len(stmt_ids) * len(student_ids) if stmt_ids else 0

        if total_possible == 0:
            continue

        # Get latest record per student+statement (within date window)
        from django.db.models import Subquery, OuterRef

        latest_qs = AssessmentRecord.objects.filter(
            student_id=OuterRef("student_id"),
            statement_id=OuterRef("statement_id"),
        )
        latest_qs = _apply_date_filter_to_records(latest_qs, date_from, date_to)
        latest_pks = latest_qs.order_by("-assessed_date", "-created_at").values("pk")[:1]

        status_counts = (
            AssessmentRecord.objects.filter(
                student_id__in=student_ids,
                statement_id__in=stmt_ids,
                pk__in=Subquery(latest_pks),
            )
            .values("status")
            .annotate(count=Count("pk"))
        )

        counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
        assessed_total = 0
        for sc in status_counts:
            counts[sc["status"]] = sc["count"]
            assessed_total += sc["count"]

        # Everything unassessed counts as NYA
        counts["NYA"] += total_possible - assessed_total

        pcts = {}
        for k, v in counts.items():
            pcts[k] = round(v / total_possible * 100, 1) if total_possible else 0

        subject_data.append({
            "subject": subject,
            "counts": counts,
            "percentages": pcts,
            "total_possible": total_possible,
            "num_students": len(student_ids),
            "num_statements": len(stmt_ids),
        })

    time_label = _time_label(date_from, date_to, time_filters)

    context = {
        "subject_data": subject_data,
        "filter_options": filter_options,
        "active_filters": active_filters,
        "num_students": len(student_ids),
        "time_label": time_label,
    }
    return render(request, "reports/subject_progress.html", context)


@login_required
def cohort_export_excel(request):
    """Export cohort report grid as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return HttpResponse("openpyxl is not installed.", status=500)

    students = Student.objects.filter(is_active=True).select_related("class_group")
    students, _ = _apply_student_filters(request, students)
    date_from, date_to, _ = _get_date_range(request)

    areas = AssessmentArea.objects.select_related("subject", "framework").prefetch_related("statements")
    areas = _apply_area_filters(request, areas)

    statements = []
    for area in areas:
        for stmt in area.statements.all():
            statements.append(stmt)

    # Build grid
    from django.db.models import Subquery, OuterRef

    grid = {}
    if students.exists() and statements:
        stmt_ids = [s.pk for s in statements]
        student_ids = list(students.values_list("pk", flat=True))

        latest_qs = AssessmentRecord.objects.filter(
            student_id=OuterRef("student_id"),
            statement_id=OuterRef("statement_id"),
        )
        latest_qs = _apply_date_filter_to_records(latest_qs, date_from, date_to)
        latest_pks = latest_qs.order_by("-assessed_date", "-created_at").values("pk")[:1]

        records = AssessmentRecord.objects.filter(
            student_id__in=student_ids,
            statement_id__in=stmt_ids,
            pk__in=Subquery(latest_pks),
        ).values("student_id", "statement_id", "status")

        for r in records:
            grid.setdefault(r["student_id"], {})[r["statement_id"]] = r["status"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cohort Report"

    fills = {
        "SEC": PatternFill(start_color="198754", end_color="198754", fill_type="solid"),
        "DEV": PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid"),
        "EME": PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
        "NYA": PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid"),
    }
    white_font = Font(color="FFFFFF", bold=True)

    # Header row
    headers = ["Student", "Class", "Pathway", "Phase"]
    for stmt in statements:
        headers.append(stmt.statement_text[:40])
    ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for student in students:
        row = [student.full_name, str(student.class_group or ""), student.get_pathway_display(), student.phase]
        s_grid = grid.get(student.pk, {})
        for stmt in statements:
            status = s_grid.get(stmt.pk, "NYA")
            row.append(status)
        ws.append(row)

        # Colour the status cells
        row_idx = ws.max_row
        for col_idx, stmt in enumerate(statements, start=5):
            cell = ws.cell(row=row_idx, column=col_idx)
            fill = fills.get(cell.value)
            if fill:
                cell.fill = fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")

    # Auto-width first 4 columns
    for col_idx in range(1, 5):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cohort_report.xlsx"'
    wb.save(response)
    return response


@login_required
def subject_progress_export_excel(request):
    """Export subject progress summary as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return HttpResponse("openpyxl is not installed.", status=500)

    students = Student.objects.filter(is_active=True)
    students, _ = _apply_student_filters(request, students)
    date_from, date_to, _ = _get_date_range(request)
    student_ids = list(students.values_list("pk", flat=True))

    subjects = Subject.objects.filter(is_active=True)
    subject_filter = request.GET.getlist("subject")
    if subject_filter:
        subjects = subjects.filter(pk__in=[int(s) for s in subject_filter])

    framework_filter = request.GET.getlist("framework")

    from django.db.models import Subquery, OuterRef

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subject Progress"

    ws.append(["Subject", "Students", "Statements", "NYA %", "Emerging %", "Developing %", "Secure %",
               "NYA #", "Emerging #", "Developing #", "Secure #"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for subject in subjects:
        areas = AssessmentArea.objects.filter(subject=subject)
        if framework_filter:
            areas = areas.filter(framework_id__in=[int(f) for f in framework_filter])

        stmt_ids = list(
            AssessmentStatement.objects.filter(area__in=areas).values_list("pk", flat=True)
        )
        total_possible = len(stmt_ids) * len(student_ids) if stmt_ids else 0
        if total_possible == 0:
            continue

        latest_qs = AssessmentRecord.objects.filter(
            student_id=OuterRef("student_id"),
            statement_id=OuterRef("statement_id"),
        )
        latest_qs = _apply_date_filter_to_records(latest_qs, date_from, date_to)
        latest_pks = latest_qs.order_by("-assessed_date", "-created_at").values("pk")[:1]

        status_counts = (
            AssessmentRecord.objects.filter(
                student_id__in=student_ids,
                statement_id__in=stmt_ids,
                pk__in=Subquery(latest_pks),
            )
            .values("status")
            .annotate(count=Count("pk"))
        )

        counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
        assessed_total = 0
        for sc in status_counts:
            counts[sc["status"]] = sc["count"]
            assessed_total += sc["count"]
        counts["NYA"] += total_possible - assessed_total

        pcts = {}
        for k, v in counts.items():
            pcts[k] = round(v / total_possible * 100, 1) if total_possible else 0

        ws.append([
            subject.name, len(student_ids), len(stmt_ids),
            pcts["NYA"], pcts["EME"], pcts["DEV"], pcts["SEC"],
            counts["NYA"], counts["EME"], counts["DEV"], counts["SEC"],
        ])

    for col_idx in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="subject_progress.xlsx"'
    wb.save(response)
    return response


# ── PDF Exports ──────────────────────────────────────────────────────

@login_required
def cohort_export_pdf(request):
    """Export cohort report grid as a PDF."""
    import weasyprint

    students = Student.objects.filter(is_active=True).select_related("class_group")
    students, _ = _apply_student_filters(request, students)
    date_from, date_to, time_filters = _get_date_range(request)

    areas = AssessmentArea.objects.select_related("subject", "framework").prefetch_related("statements")
    areas = _apply_area_filters(request, areas)

    statements = []
    for area in areas:
        for stmt in area.statements.all():
            statements.append(stmt)

    # Build grid
    from django.db.models import Subquery, OuterRef

    grid = {}
    if students.exists() and statements:
        stmt_ids = [s.pk for s in statements]
        student_ids = list(students.values_list("pk", flat=True))

        latest_qs = AssessmentRecord.objects.filter(
            student_id=OuterRef("student_id"),
            statement_id=OuterRef("statement_id"),
        )
        latest_qs = _apply_date_filter_to_records(latest_qs, date_from, date_to)
        latest_pks = latest_qs.order_by("-assessed_date", "-created_at").values("pk")[:1]

        records = AssessmentRecord.objects.filter(
            student_id__in=student_ids,
            statement_id__in=stmt_ids,
            pk__in=Subquery(latest_pks),
        ).values("student_id", "statement_id", "status")

        for r in records:
            grid.setdefault(r["student_id"], {})[r["statement_id"]] = r["status"]

    # Pre-build rows for template (avoids dict-key lookup in Django templates)
    rows = []
    for student in students:
        s_grid = grid.get(student.pk, {})
        counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
        cells = []
        for stmt in statements:
            status = s_grid.get(stmt.pk, "NYA")
            counts[status] = counts.get(status, 0) + 1
            cells.append(status)
        rows.append({
            "name": student.full_name,
            "class_name": str(student.class_group or "—"),
            "sec": counts["SEC"],
            "dev": counts["DEV"],
            "eme": counts["EME"],
            "nya": counts["NYA"],
            "cells": cells,
        })

    time_label = _time_label(date_from, date_to, time_filters)
    stmt_headers = [s.statement_text[:30] for s in statements]

    context = {
        "rows": rows,
        "stmt_headers": stmt_headers,
        "time_label": time_label,
        "today": timezone.now().date(),
    }
    html_string = render_to_string("pdf/cohort_report.html", context)
    pdf = weasyprint.HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cohort_report.pdf"'
    return response


@login_required
def subject_progress_export_pdf(request):
    """Export subject progress summary as a PDF."""
    import weasyprint

    students = Student.objects.filter(is_active=True)
    students, _ = _apply_student_filters(request, students)
    date_from, date_to, time_filters = _get_date_range(request)
    student_ids = list(students.values_list("pk", flat=True))

    subjects = Subject.objects.filter(is_active=True)
    subject_filter = request.GET.getlist("subject")
    if subject_filter:
        subjects = subjects.filter(pk__in=[int(s) for s in subject_filter])

    framework_filter = request.GET.getlist("framework")

    from django.db.models import Subquery, OuterRef

    subject_data = []
    for subject in subjects:
        areas_qs = AssessmentArea.objects.filter(subject=subject)
        if framework_filter:
            areas_qs = areas_qs.filter(framework_id__in=[int(f) for f in framework_filter])

        stmt_ids = list(
            AssessmentStatement.objects.filter(area__in=areas_qs).values_list("pk", flat=True)
        )
        total_possible = len(stmt_ids) * len(student_ids) if stmt_ids else 0
        if total_possible == 0:
            continue

        latest_qs = AssessmentRecord.objects.filter(
            student_id=OuterRef("student_id"),
            statement_id=OuterRef("statement_id"),
        )
        latest_qs = _apply_date_filter_to_records(latest_qs, date_from, date_to)
        latest_pks = latest_qs.order_by("-assessed_date", "-created_at").values("pk")[:1]

        status_counts = (
            AssessmentRecord.objects.filter(
                student_id__in=student_ids,
                statement_id__in=stmt_ids,
                pk__in=Subquery(latest_pks),
            )
            .values("status")
            .annotate(count=Count("pk"))
        )

        counts = {"SEC": 0, "DEV": 0, "EME": 0, "NYA": 0}
        assessed_total = 0
        for sc in status_counts:
            counts[sc["status"]] = sc["count"]
            assessed_total += sc["count"]
        counts["NYA"] += total_possible - assessed_total

        pcts = {}
        for k, v in counts.items():
            pcts[k] = round(v / total_possible * 100, 1) if total_possible else 0

        subject_data.append({
            "subject": subject,
            "counts": counts,
            "percentages": pcts,
            "total_possible": total_possible,
            "num_students": len(student_ids),
            "num_statements": len(stmt_ids),
        })

    time_label = _time_label(date_from, date_to, time_filters)

    context = {
        "subject_data": subject_data,
        "num_students": len(student_ids),
        "time_label": time_label,
        "today": timezone.now().date(),
    }
    html_string = render_to_string("pdf/subject_progress.html", context)
    pdf = weasyprint.HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="subject_progress.pdf"'
    return response
